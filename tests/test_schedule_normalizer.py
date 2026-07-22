from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / ".agents/skills/xdf-normalize-teacher-schedule/scripts/normalize_schedule.py"
RULE_MANAGER = ROOT / ".agents/skills/xdf-normalize-teacher-schedule/scripts/manage_schedule_rules.py"
SPEC = importlib.util.spec_from_file_location("normalize_schedule", SCRIPT)
assert SPEC and SPEC.loader
NORMALIZE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(NORMALIZE)

INPUT = ROOT / "evals/fixtures/synthetic-schedule-input.xlsx"
MANUAL = ROOT / "evals/fixtures/synthetic-manual-schedule.xlsx"
CONFIG = ROOT / "config/teacher-schedule.json"


class ScheduleNormalizerTest(unittest.TestCase):
    def run_case(
        self,
        input_path: Path,
        output_dir: Path,
        *,
        campus: str | list[str] | None = "合成校区",
        campus_group: str | None = None,
        manual_template: Path | None = None,
        manual_templates: list[Path] | None = None,
        config_path: Path = CONFIG,
        template_policy: str = "rule_first",
        rules_overlay: Path | None = None,
    ):
        return NORMALIZE.run_schedule(
            input_path,
            config_path,
            output_dir,
            campus=campus,
            campus_group=campus_group,
            data_as_of="2026-07-22",
            manual_template=manual_template,
            manual_templates=manual_templates,
            template_policy=template_policy,
            rules_overlay=rules_overlay,
        )

    def mutate_workbook(self, source: Path, folder: Path, name: str, mutator) -> Path:
        target = folder / name
        workbook = load_workbook(source)
        mutator(workbook)
        workbook.save(target)
        workbook.close()
        return target

    def test_baseline_rules_conflicts_colors_and_formulas(self):
        with tempfile.TemporaryDirectory() as directory:
            result = self.run_case(INPUT, Path(directory) / "out")
            self.assertEqual("needs_confirmation", result["status"])

            issue_codes = {issue["code"] for issue in result["issues"]}
            self.assertTrue(
                {
                    "NONSTANDARD_SESSIONS",
                    "DIFFICULTY_CONFLICT",
                    "EXCLUDED_GRADE",
                    "EXCLUDED_CLASS_NAME",
                    "CLOSED_CLASS",
                    "MISSING_TEACHER",
                    "TEACHER_TIME_CONFLICT_SELECTED",
                    "TEACHER_TIME_CONFLICT_TIE",
                }.issubset(issue_codes)
            )
            exclusion_messages = [issue["message"] for issue in result["issues"] if issue["severity"] == "排除"]
            self.assertTrue(any("命中排除词：补课" in message for message in exclusion_messages))
            self.assertTrue(any("班级状态“取消”" in message for message in exclusion_messages))

            selected = {record["class_code"]: record for record in result["selected"]}
            self.assertIn("ST3CS260008", selected)
            self.assertNotIn("ST3BS260007", selected)
            self.assertNotIn("SW5AS260009", selected)
            self.assertNotIn("SW5BS260010", selected)
            self.assertEqual("B", selected["SW4BS260011"]["difficulty"])
            self.assertEqual(-1, selected["ST1AF260105"]["maximum"] - selected["ST1AF260105"]["current_students"])

            workbook = load_workbook(result["outputs"]["workbook"], data_only=False)
            try:
                self.assertIn("周三18点", [cell.value for cell in workbook["秋季课表"][1]])
                self.assertEqual("=I4-J4", workbook["容量明细"]["K4"].value)
                self.assertEqual("=G4-H4", workbook["容量汇总"]["I4"].value)
                self.assertEqual("=E4-F4", workbook["容量总览"]["G4"].value)
                self.assertEqual("难度/班型", workbook["容量明细"]["D3"].value)
                self.assertIn("排除记录", workbook.sheetnames)
                self.assertIn("模板对比", workbook.sheetnames)
                self.assertNotIn("排除", {cell.value for cell in workbook["校验问题"]["A"][3:]})

                summer = workbook["暑假课表"]
                cells = {cell.value: cell for row in summer.iter_rows() for cell in row if cell.value}
                self.assertEqual("001D4ED8", cells["11人\nST3AS260001"].font.color.rgb)
                self.assertEqual("00157A3D", cells["15人\nSW4BS260002"].font.color.rgb)
                self.assertEqual("00B42318", cells["16人\nBX2AS260003"].font.color.rgb)
            finally:
                workbook.close()

    def test_planned_classes_and_low_grade_nl_rules(self):
        with tempfile.TemporaryDirectory() as directory:
            result = self.run_case(INPUT, Path(directory) / "out", manual_template=MANUAL)
            selected = {record["class_code"]: record for record in result["selected"]}
            self.assertEqual(("基础", 16, 0), tuple(selected["新班NL1"][key] for key in ("difficulty", "maximum", "current_students")))
            self.assertEqual(("拔高", 16, 0), tuple(selected["新班NL2A"][key] for key in ("difficulty", "maximum", "current_students")))
            self.assertEqual(("B", 20, 0), tuple(selected["新班SW3B"][key] for key in ("difficulty", "maximum", "current_students")))
            issue_codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("INVALID_NL_PLANNED_SUFFIX", issue_codes)
            self.assertIn("STALE_TEMPLATE_CLASS", issue_codes)
            self.assertIn("PLANNED_CLASS_MISSING_TEACHER", issue_codes)
            self.assertIn("PLANNED_SUBJECT_CONFLICT", issue_codes)

            review = json.loads(Path(result["outputs"]["review"]).read_text(encoding="utf-8"))
            self.assertTrue(review["requires_human_confirmation"])
            self.assertFalse(review["auto_apply"])
            self.assertIn("关班", review["human_followup_prompt"])
            self.assertIn("更换教师", review["human_followup_prompt"])

    def test_manual_move_of_official_class_is_detected_but_not_applied(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def move_official_class(workbook):
                workbook["暑假课表"]["E4"] = "11人\nST3AS260001"

            manual = self.mutate_workbook(MANUAL, root, "moved-official.xlsx", move_official_class)
            result = self.run_case(INPUT, root / "out", manual_template=manual)
            issue_codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("MANUAL_OFFICIAL_CLASS_MOVE", issue_codes)
            official = next(record for record in result["selected"] if record["class_code"] == "ST3AS260001")
            self.assertEqual(("合成教师甲", "一轮08点"), (official["teacher"], official["time_key"]))

    def test_template_first_applies_valid_manual_move_and_removal(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def move_official_class(workbook):
                workbook["暑假课表"]["E4"] = "11人\nST3AS260001"

            manual = self.mutate_workbook(MANUAL, root, "moved-official.xlsx", move_official_class)
            result = self.run_case(
                INPUT,
                root / "out",
                manual_template=manual,
                template_policy="template_first",
            )
            selected = {record["class_code"]: record for record in result["selected"]}
            self.assertEqual(("合成计划教师丙", "二轮10点"), (selected["ST3AS260001"]["teacher"], selected["ST3AS260001"]["time_key"]))
            self.assertNotIn("SW4BS260002", selected)
            issue_codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("TEMPLATE_OVERRIDE_APPLIED", issue_codes)
            self.assertIn("TEMPLATE_OVERRIDE_SOURCE_REMOVED", issue_codes)
            self.assertNotIn("MANUAL_OFFICIAL_CLASS_MOVE", issue_codes)
            review = json.loads(Path(result["outputs"]["review"]).read_text(encoding="utf-8"))
            self.assertEqual("template_first", review["template_policy"])
            self.assertTrue(review["template_changes_applied"])

    def test_template_first_rejects_duplicate_positions_and_ineligible_source_class(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def make_adversarial_template(workbook):
                sheet = workbook["暑假课表"]
                sheet["D4"] = "11人\nST4AS260006"
                sheet["E4"] = "11人\nST3AS260001"
                sheet["F4"] = "11人\nST3AS260001"

            manual = self.mutate_workbook(MANUAL, root, "adversarial-template.xlsx", make_adversarial_template)
            result = self.run_case(INPUT, root / "out", manual_template=manual, template_policy="template_first")
            selected_codes = {record["class_code"] for record in result["selected"]}
            self.assertNotIn("ST3AS260001", selected_codes)
            self.assertNotIn("ST4AS260006", selected_codes)
            issue_codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("DUPLICATE_TEMPLATE_CLASS", issue_codes)
            self.assertIn("TEMPLATE_CLASS_NOT_ELIGIBLE", issue_codes)

    def test_multiple_teachers_use_unique_matching_primary_teacher(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def add_primary_teacher(workbook):
                sheet = workbook["暑假明细"]
                headers = [cell.value for cell in sheet[1]]
                teacher_column = headers.index("授课教师") + 1
                primary_column = sheet.max_column + 1
                sheet.cell(1, primary_column, "主带课老师")
                sheet.cell(2, teacher_column, "合成教师甲(1001),合成教师乙(1002)")
                sheet.cell(2, primary_column, "合成教师乙(1002)")

            path = self.mutate_workbook(INPUT, root, "primary-teacher.xlsx", add_primary_teacher)
            result = self.run_case(path, root / "out")
            record = next(record for record in result["selected"] if record["class_code"] == "ST3AS260001")
            self.assertEqual("合成教师乙", record["teacher"])
            self.assertEqual("主带课老师", record["teacher_basis"])
            self.assertIn("PRIMARY_TEACHER_SELECTED", {issue["code"] for issue in result["issues"]})

            def mismatch_primary_teacher(workbook):
                sheet = workbook["暑假明细"]
                headers = [cell.value for cell in sheet[1]]
                teacher_column = headers.index("授课教师") + 1
                primary_column = sheet.max_column + 1
                sheet.cell(1, primary_column, "主带课老师")
                sheet.cell(2, teacher_column, "合成教师甲(1001),合成教师乙(1002)")
                sheet.cell(2, primary_column, "合成教师丙(1003)")

            mismatch = self.mutate_workbook(INPUT, root, "mismatch-primary.xlsx", mismatch_primary_teacher)
            mismatch_result = self.run_case(mismatch, root / "mismatch-out")
            self.assertNotIn("ST3AS260001", {record["class_code"] for record in mismatch_result["selected"]})
            self.assertIn("PRIMARY_TEACHER_MISMATCH", {issue["code"] for issue in mismatch_result["issues"]})

    def test_unknown_campus_blocks_and_only_emits_validation_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            result = self.run_case(INPUT, Path(directory) / "out", campus="不存在校区")
            self.assertEqual("blocked", result["status"])
            self.assertIn("NO_CAMPUS_RECORDS", {issue["code"] for issue in result["issues"]})
            workbook = load_workbook(result["outputs"]["workbook"])
            try:
                self.assertEqual(["运行摘要", "校验问题"], workbook.sheetnames)
            finally:
                workbook.close()

    def test_missing_column_and_duplicate_class_code_block(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            missing = self.mutate_workbook(INPUT, root, "missing.xlsx", lambda wb: setattr(wb["暑假明细"]["A1"], "value", "未知字段"))
            missing_result = self.run_case(missing, root / "missing-out")
            self.assertEqual("blocked", missing_result["status"])
            self.assertIn("MISSING_COLUMN", {issue["code"] for issue in missing_result["issues"]})

            def duplicate(workbook):
                workbook["秋季明细"]["C2"] = workbook["暑假明细"]["C2"].value

            duplicate_path = self.mutate_workbook(INPUT, root, "duplicate.xlsx", duplicate)
            duplicate_result = self.run_case(duplicate_path, root / "duplicate-out")
            self.assertEqual("blocked", duplicate_result["status"])
            self.assertIn("DUPLICATE_CLASS_CODE", {issue["code"] for issue in duplicate_result["issues"]})

    def test_input_formula_external_link_and_manual_formula_block(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            formula_path = self.mutate_workbook(INPUT, root, "formula.xlsx", lambda wb: setattr(wb["暑假明细"]["N2"], "value", "=1+1"))
            formula_result = self.run_case(formula_path, root / "formula-out")
            self.assertEqual("blocked", formula_result["status"])
            self.assertTrue({"EXCEL_FORMULA", "FORMULA_INJECTION"} & {issue["code"] for issue in formula_result["issues"]})

            external_path = root / "external.xlsx"
            with zipfile.ZipFile(INPUT) as source, zipfile.ZipFile(external_path, "w") as destination:
                for item in source.infolist():
                    destination.writestr(item, source.read(item.filename))
                destination.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")
            external_result = self.run_case(external_path, root / "external-out")
            self.assertEqual("blocked", external_result["status"])
            self.assertIn("EXTERNAL_LINK", {issue["code"] for issue in external_result["issues"]})

            hyperlink_path = self.mutate_workbook(
                INPUT,
                root,
                "hyperlink.xlsx",
                lambda wb: setattr(wb["暑假明细"]["N2"], "hyperlink", "https://example.invalid"),
            )
            hyperlink_result = self.run_case(hyperlink_path, root / "hyperlink-out")
            self.assertEqual("blocked", hyperlink_result["status"])
            self.assertIn("EXTERNAL_LINK", {issue["code"] for issue in hyperlink_result["issues"]})

            malicious_manual = self.mutate_workbook(MANUAL, root, "manual-formula.xlsx", lambda wb: setattr(wb["暑假课表"]["D2"], "value", "=HYPERLINK(\"https://example.invalid\",\"x\")"))
            manual_result = self.run_case(INPUT, root / "manual-out", manual_template=malicious_manual)
            self.assertEqual("blocked", manual_result["status"])
            self.assertIn("UNEXPECTED_MANUAL_FORMULA", {issue["code"] for issue in manual_result["issues"]})

    def test_source_capacity_conflict_uses_versioned_rule(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            def change_two_capacities(workbook):
                workbook["暑假明细"]["F2"] = 99
                workbook["暑假明细"]["F8"] = 99

            path = self.mutate_workbook(INPUT, root, "capacity-conflict.xlsx", change_two_capacities)
            result = self.run_case(path, root / "out")
            selected = {record["class_code"]: record for record in result["selected"]}
            self.assertEqual(20, selected["ST3AS260001"]["maximum"])
            conflicts = [issue for issue in result["issues"] if issue["code"] == "SOURCE_MAX_CONFLICT"]
            self.assertEqual(1, len(conflicts))
            self.assertEqual(2, conflicts[0]["affected_count"])

    def test_configured_campus_group_and_repeatable_campuses(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = json.loads(CONFIG.read_text(encoding="utf-8"))
            config["business"]["campus_groups"]["合成管理单元"] = ["合成校区", "其他合成校区"]
            config_path = root / "group-config.json"
            config_path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")

            direct = self.run_case(INPUT, root / "direct", campus=["合成校区", "其他合成校区"])
            grouped = self.run_case(INPUT, root / "grouped", campus=None, campus_group="合成管理单元", config_path=config_path)
            for result in (direct, grouped):
                self.assertIn("ST6AS260015", {record["class_code"] for record in result["selected"]})
                self.assertEqual(["合成校区", "其他合成校区"], result["campuses"])
            self.assertEqual("合成管理单元", grouped["campus"])

    def test_hybrid_summer_schedule_uses_intensive_daily_segment(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def make_hybrid(workbook):
                sheet = workbook["暑假明细"]
                sheet["G2"] = "2026-06-13"
                sheet["H2"] = "2026.6.13-2026.6.27每周三18:00-20:00;2026.7.1-2026.7.7每天13:50-15:50"
                sheet["M2"] = ""

            path = self.mutate_workbook(INPUT, root, "hybrid.xlsx", make_hybrid)
            result = self.run_case(path, root / "out")
            record = next(record for record in result["selected"] if record["class_code"] == "ST3AS260001")
            self.assertEqual("一轮13点", record["time_key"])
            self.assertIn("HYBRID_SCHEDULE_RESOLVED", {issue["code"] for issue in result["issues"]})

            def make_ambiguous(workbook):
                sheet = workbook["暑假明细"]
                sheet["G2"] = "2026-06-13"
                sheet["H2"] = "2026.7.1-2026.7.7每天13:50-15:50;2026.7.21-2026.7.27每天10:40-12:40"
                sheet["M2"] = ""

            ambiguous = self.mutate_workbook(INPUT, root, "ambiguous-hybrid.xlsx", make_ambiguous)
            ambiguous_result = self.run_case(ambiguous, root / "ambiguous-out")
            self.assertNotIn("ST3AS260001", {record["class_code"] for record in ambiguous_result["selected"]})
            self.assertIn("AMBIGUOUS_INTENSIVE_SCHEDULE", {issue["code"] for issue in ambiguous_result["issues"]})

    def test_multiple_manual_templates_three_way_compare_and_ignore_unsupported_codes(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def summer_only(workbook):
                workbook.remove(workbook["秋季课表"])
                workbook["暑假课表"]["E4"] = "3人\nKC1AS260999"

            def fall_only(workbook):
                workbook.remove(workbook["暑假课表"])

            summer = self.mutate_workbook(MANUAL, root, "summer-only.xlsx", summer_only)
            fall = self.mutate_workbook(MANUAL, root, "fall-only.xlsx", fall_only)
            result = self.run_case(INPUT, root / "out", manual_templates=[summer, fall])
            review = json.loads(Path(result["outputs"]["review"]).read_text(encoding="utf-8"))
            self.assertEqual(2, review["manual_template_count"])
            self.assertGreater(review["manual_comparison_counts"]["source_only"], 0)
            self.assertIn("STALE_TEMPLATE_CLASS", {issue["code"] for issue in result["issues"]})
            self.assertNotIn("KC1AS260999", {issue["class_code"] for issue in result["issues"]})

    def test_empty_conflict_sheet_has_explicit_message(self):
        with tempfile.TemporaryDirectory() as directory:
            result = self.run_case(INPUT, Path(directory) / "out", campus="其他合成校区")
            workbook = load_workbook(result["outputs"]["workbook"], data_only=False)
            try:
                self.assertEqual("本次未发现教师时段冲突", workbook["排课冲突"]["A4"].value)
            finally:
                workbook.close()

    def test_invalid_capacity_and_low_contrast_config_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            base = json.loads(CONFIG.read_text(encoding="utf-8"))
            base["business"]["capacity_rules"]["思维"]["1-2"] = {"minimum": 17, "maximum": 16}
            invalid_capacity = root / "invalid-capacity.json"
            invalid_capacity.write_text(json.dumps(base, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "开班线/最大班容无效"):
                NORMALIZE.load_config(invalid_capacity)

            base = json.loads(CONFIG.read_text(encoding="utf-8"))
            base["colors"]["open_not_full"] = "FFFFFF"
            low_contrast = root / "low-contrast.json"
            low_contrast.write_text(json.dumps(base, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "对比度"):
                NORMALIZE.load_config(low_contrast)

            base = json.loads(CONFIG.read_text(encoding="utf-8"))
            base["business"]["campus_groups"]["重复校区"] = ["合成校区", "合成校区"]
            invalid_group = root / "invalid-group.json"
            invalid_group.write_text(json.dumps(base, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "无重复"):
                NORMALIZE.load_config(invalid_group)

    def test_local_rule_overlay_is_reversible_and_cannot_touch_hard_paths(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            overlay = root / "active-overrides.json"
            patch = root / "patch.json"
            patch.write_text(
                json.dumps(
                    {"business": {"capacity_rules": {"思维": {"1-2": {"minimum": 8, "maximum": 17}}}}},
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            original = CONFIG.read_bytes()
            apply_result = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(RULE_MANAGER),
                    "--base",
                    str(CONFIG),
                    "--overlay",
                    str(overlay),
                    "apply",
                    "--patch",
                    str(patch),
                    "--request",
                    "思维一二年级最大班容调到17",
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, apply_result.returncode, apply_result.stderr)
            self.assertEqual(original, CONFIG.read_bytes())
            effective = NORMALIZE.load_config(CONFIG, overlay)
            self.assertEqual(17, effective["business"]["capacity_rules"]["思维"]["1-2"]["maximum"])
            self.assertEqual("local_override", effective["_rules"]["mode"])

            restore_result = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(RULE_MANAGER),
                    "--base",
                    str(CONFIG),
                    "--overlay",
                    str(overlay),
                    "restore",
                    "--request",
                    "回到初始规则",
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, restore_result.returncode, restore_result.stderr)
            restored = NORMALIZE.load_config(CONFIG, overlay)
            self.assertEqual(16, restored["business"]["capacity_rules"]["思维"]["1-2"]["maximum"])
            self.assertEqual("initial", restored["_rules"]["mode"])
            self.assertEqual(original, CONFIG.read_bytes())

            disallowed = root / "disallowed.json"
            disallowed.write_text(json.dumps({"source": {"seasons": {}}}, ensure_ascii=False), encoding="utf-8")
            rejected = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(RULE_MANAGER),
                    "--base",
                    str(CONFIG),
                    "--overlay",
                    str(overlay),
                    "apply",
                    "--patch",
                    str(disallowed),
                    "--request",
                    "修改源表定义",
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(1, rejected.returncode)
            self.assertIn("不允许覆盖规则路径", rejected.stderr)

            injection = root / "injection.json"
            injection.write_text(
                json.dumps({"business": {"weekly_default_columns": ["=HYPERLINK(\"https://example.invalid\",\"x\")"]}}),
                encoding="utf-8",
            )
            injected = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(RULE_MANAGER),
                    "--base",
                    str(CONFIG),
                    "--overlay",
                    str(overlay),
                    "apply",
                    "--patch",
                    str(injection),
                    "--request",
                    "修改默认时段",
                ],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(1, injected.returncode)
            self.assertIn("公式注入", injected.stderr)

    def test_stale_rule_overlay_and_invalid_template_policy_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            base = root / "base.json"
            base.write_bytes(CONFIG.read_bytes())
            overlay = root / "overlay.json"
            overlay.write_text(
                json.dumps(
                    {
                        "schema_version": "1.0.0",
                        "base_config_sha256": NORMALIZE.base_config_sha256(base),
                        "overrides": {},
                        "history": [],
                    }
                ),
                encoding="utf-8",
            )
            config = json.loads(base.read_text(encoding="utf-8"))
            config["schema_version"] = "1.2.1"
            base.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "初始规则已变化"):
                NORMALIZE.load_config(base, overlay)
            with self.assertRaisesRegex(ValueError, "template_policy"):
                self.run_case(INPUT, root / "out", template_policy="unknown")


if __name__ == "__main__":
    unittest.main()
