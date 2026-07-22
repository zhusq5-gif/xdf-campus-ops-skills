#!/usr/bin/env python3
"""Deterministically convert class-detail exports into teacher schedule matrices."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLOCKING = "阻断"
PENDING = "待确认"
WARNING = "警告"
INFO = "信息"
EXCLUDED = "排除"
FORMULA_PREFIXES = ("=", "+", "-", "@")
WEEKDAY_MAP = {
    "一": "周一",
    "二": "周二",
    "三": "周三",
    "四": "周四",
    "五": "周五",
    "六": "周六",
    "日": "周日",
    "天": "周日",
    "1": "周一",
    "2": "周二",
    "3": "周三",
    "4": "周四",
    "5": "周五",
    "6": "周六",
    "7": "周日",
}
CHINESE_GRADES = {"一年级": 1, "二年级": 2, "三年级": 3, "四年级": 4, "五年级": 5, "六年级": 6}
REQUIRED_SOURCE_FIELDS = {
    "subject",
    "campus",
    "class_code",
    "class_name",
    "current_students",
    "source_max_students",
    "start_date",
    "schedule_text",
    "teacher",
    "class_sessions",
}
REQUIRED_COLORS = {
    "header",
    "header_text",
    "subject_fill",
    "teacher_fill",
    "border",
    "weekday_fill",
    "saturday_fill",
    "sunday_fill",
    "round_fills",
    "below_minimum",
    "open_not_full",
    "full_or_oversold",
    "text",
}
ALLOWED_RULE_OVERRIDE_ROOTS = {
    ("business", "campus_groups"),
    ("business", "expected_sessions"),
    ("business", "capacity_rules"),
    ("business", "time_slots"),
    ("business", "rounds"),
    ("business", "summer_round_dates"),
    ("business", "summer_intensive_markers"),
    ("business", "weekday_order"),
    ("business", "weekly_default_columns"),
    ("business", "excluded_grade_labels"),
    ("business", "excluded_name_keywords"),
    ("business", "closed_status_values"),
    ("colors",),
}


def add_issue(
    issues: list[dict[str, Any]],
    severity: str,
    code: str,
    message: str,
    *,
    season: str = "",
    source_ref: str = "",
    class_code: str = "",
    teacher: str = "",
    time_key: str = "",
    action: str = "",
    affected_count: int = 1,
) -> None:
    issues.append(
        {
            "severity": severity,
            "code": code,
            "season": season,
            "source_ref": source_ref,
            "class_code": class_code,
            "teacher": teacher,
            "time_key": time_key,
            "message": message,
            "action": action,
            "affected_count": affected_count,
        }
    )


def validate_config(config: dict[str, Any]) -> dict[str, Any]:
    required = {"schema_version", "output_version", "source", "business", "colors"}
    missing = required - set(config)
    if missing:
        raise ValueError(f"配置缺少字段: {', '.join(sorted(missing))}")
    if not config["source"].get("seasons"):
        raise ValueError("配置必须包含至少一个季节源表")
    business = config["business"]
    campus_groups = business.get("campus_groups", {})
    if not isinstance(campus_groups, dict):
        raise ValueError("campus_groups 必须是管理单元到源校区列表的映射")
    for label, campuses in campus_groups.items():
        if (
            not isinstance(label, str)
            or not label.strip()
            or not isinstance(campuses, list)
            or not campuses
            or len(campuses) != len(set(campuses))
            or any(not isinstance(campus, str) or not campus.strip() for campus in campuses)
        ):
            raise ValueError("每个 campus_groups 项必须包含非空、无重复的源校区名称列表")
    for season, spec in config["source"]["seasons"].items():
        if spec.get("mode") not in {"round", "weekly"}:
            raise ValueError(f"{season} 的 mode 必须是 round 或 weekly")
        missing_fields = REQUIRED_SOURCE_FIELDS - set(spec.get("columns", {}))
        if missing_fields:
            raise ValueError(f"{season} 的字段映射不完整: {', '.join(sorted(missing_fields))}")
        if season not in business.get("expected_sessions", {}):
            raise ValueError(f"缺少 {season} 的标准课次规则")
        for grade in business["valid_grades"]:
            sessions = range_value(business["expected_sessions"][season], grade)
            if type(sessions) is not int or sessions <= 0:
                raise ValueError(f"{season} {grade} 年级的标准课次无效")
    for subject in set(business.get("subject_by_prefix", {}).values()):
        rules = business.get("capacity_rules", {}).get(subject)
        if not rules:
            raise ValueError(f"缺少科目 {subject} 的容量规则")
        for grade in business["valid_grades"]:
            capacity = range_value(rules, grade)
            if not isinstance(capacity, dict):
                raise ValueError(f"缺少科目 {subject} {grade} 年级的容量规则")
            minimum = capacity.get("minimum")
            maximum = capacity.get("maximum")
            if type(minimum) is not int or type(maximum) is not int or minimum < 0 or maximum < minimum:
                raise ValueError(f"科目 {subject} {grade} 年级的开班线/最大班容无效")
    subjects = set(business["subject_by_prefix"].values())
    if len(business.get("subject_order", [])) != len(subjects) or set(business["subject_order"]) != subjects:
        raise ValueError("subject_order 必须且只能包含全部支持科目")
    if business.get("classification_priority") != "class_code":
        raise ValueError("当前版本只支持班号优先的班型判定")
    low_grade_nl = business.get("low_grade_nl", {})
    if (
        low_grade_nl.get("grades") != [1, 2]
        or low_grade_nl.get("advanced_suffix") != "A"
        or low_grade_nl.get("planned_base_suffix") != ""
        or not low_grade_nl.get("base_label")
        or not low_grade_nl.get("advanced_label")
    ):
        raise ValueError("低年级 NL 规则必须明确为 NL1/NL2 基础、A 后缀拔高")
    conflict_policy = business.get("conflict_policy", {})
    if conflict_policy != {"multiple_classes": "unique_max_current", "highest_tie": "manual_confirmation"}:
        raise ValueError("当前版本只支持唯一最高人数临时入表、并列人工确认的冲突策略")
    if REQUIRED_COLORS - set(config["colors"]):
        raise ValueError("配色字段不完整")
    flat_colors = [value for key, value in config["colors"].items() if key != "round_fills"] + config["colors"]["round_fills"]
    if any(not isinstance(value, str) or not re.fullmatch(r"[0-9A-Fa-f]{6}", value) for value in flat_colors):
        raise ValueError("颜色必须是不带 # 的 6 位十六进制值")
    status_colors = [config["colors"][key] for key in ("below_minimum", "open_not_full", "full_or_oversold")]
    time_fills = [config["colors"][key] for key in ("weekday_fill", "saturday_fill", "sunday_fill")] + config["colors"]["round_fills"]
    if any(contrast_ratio(font, fill) < 4.5 for font in status_colors for fill in time_fills):
        raise ValueError("状态字与时间背景的对比度必须至少为 4.5:1")
    if contrast_ratio(config["colors"]["header_text"], config["colors"]["header"]) < 4.5:
        raise ValueError("表头文字与背景的对比度必须至少为 4.5:1")
    time_slots = business.get("time_slots", [])
    if not isinstance(time_slots, list) or not time_slots:
        raise ValueError("time_slots 必须是非空列表")
    occupied_hours: set[int] = set()
    for slot in time_slots:
        if not isinstance(slot, dict) or set(slot) != {"label", "minimum_hour", "maximum_hour"}:
            raise ValueError("每个 time_slots 项必须且只能包含 label、minimum_hour、maximum_hour")
        if not isinstance(slot["label"], str) or not slot["label"].strip():
            raise ValueError("时段 label 必须是非空字符串")
        minimum_hour = slot["minimum_hour"]
        maximum_hour = slot["maximum_hour"]
        if type(minimum_hour) is not int or type(maximum_hour) is not int or not 0 <= minimum_hour <= maximum_hour <= 23:
            raise ValueError("时段小时范围必须为 0-23 内的有效整数")
        hours = set(range(minimum_hour, maximum_hour + 1))
        if occupied_hours & hours:
            raise ValueError("time_slots 的小时范围不得重叠")
        occupied_hours.update(hours)
    rounds = business.get("rounds", [])
    if (
        not isinstance(rounds, list)
        or not rounds
        or len(rounds) != len(set(rounds))
        or any(not isinstance(item, str) or not item.strip() for item in rounds)
        or len(rounds) > len(config["colors"]["round_fills"])
    ):
        raise ValueError("rounds 必须是非空、无重复且具有对应配色的字符串列表")
    weekday_order = business.get("weekday_order", [])
    if (
        not isinstance(weekday_order, list)
        or not weekday_order
        or len(weekday_order) != len(set(weekday_order))
        or any(not isinstance(item, str) or not item.strip() for item in weekday_order)
    ):
        raise ValueError("weekday_order 必须是非空、无重复的字符串列表")
    try:
        re.compile(business["planned_class_pattern"])
    except (KeyError, re.error) as exc:
        raise ValueError("计划新班格式规则无效") from exc
    markers = business.get("summer_intensive_markers", [])
    if not isinstance(markers, list) or not markers or any(not isinstance(marker, str) or not marker for marker in markers):
        raise ValueError("summer_intensive_markers 必须是非空字符串列表")
    return config


def validate_rule_overrides(overrides: Any) -> dict[str, Any]:
    if not isinstance(overrides, dict):
        raise ValueError("规则覆盖 overrides 必须是 JSON 对象")
    for top_key, value in overrides.items():
        if top_key == "colors":
            if not isinstance(value, dict):
                raise ValueError("colors 覆盖必须是 JSON 对象")
            continue
        if top_key != "business" or not isinstance(value, dict):
            raise ValueError(f"不允许覆盖规则路径: {top_key}")
        for business_key in value:
            if ("business", business_key) not in ALLOWED_RULE_OVERRIDE_ROOTS:
                raise ValueError(f"不允许覆盖规则路径: business.{business_key}")
    pending = [overrides]
    while pending:
        value = pending.pop()
        if isinstance(value, dict):
            pending.extend(value.values())
        elif isinstance(value, list):
            pending.extend(value)
        elif isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
            raise ValueError("规则覆盖不得包含疑似公式注入文本")
    return overrides


def deep_merge(base: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    merged = copy.deepcopy(base)
    for key, value in overrides.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = deep_merge(merged[key], value)
        else:
            merged[key] = copy.deepcopy(value)
    return merged


def base_config_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_config(path: Path, rules_overlay: Path | None = None) -> dict[str, Any]:
    base_config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(base_config)
    config = base_config
    rules_meta: dict[str, Any] = {
        "mode": "initial",
        "base_config_sha256": base_config_sha256(path),
        "override_count": 0,
    }
    if rules_overlay is not None:
        overlay = json.loads(rules_overlay.read_text(encoding="utf-8"))
        if overlay.get("schema_version") != "1.0.0":
            raise ValueError("规则覆盖文件 schema_version 必须为 1.0.0")
        if overlay.get("base_config_sha256") != rules_meta["base_config_sha256"]:
            raise ValueError("初始规则已变化，请重新确认本地覆盖规则")
        overrides = validate_rule_overrides(overlay.get("overrides", {}))
        history = overlay.get("history", [])
        if not isinstance(history, list):
            raise ValueError("规则覆盖 history 必须是列表")
        config = deep_merge(base_config, overrides)
        validate_config(config)
        rules_meta = {
            "mode": "local_override" if overrides else "initial",
            "base_config_sha256": rules_meta["base_config_sha256"],
            "override_count": len(flatten_config(overrides)),
        }
    config = copy.deepcopy(config)
    config["_rules"] = rules_meta
    return config


def contrast_ratio(first: str, second: str) -> float:
    def luminance(color: str) -> float:
        channels = [int(color[index:index + 2], 16) / 255 for index in (0, 2, 4)]
        linear = [value / 12.92 if value <= 0.04045 else ((value + 0.055) / 1.055) ** 2.4 for value in channels]
        return 0.2126 * linear[0] + 0.7152 * linear[1] + 0.0722 * linear[2]

    light, dark = sorted((luminance(first), luminance(second)), reverse=True)
    return (light + 0.05) / (dark + 0.05)


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def parse_integer(value: Any) -> int | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or not number.is_integer():
        return None
    return int(number)


def clean_teacher_names(value: Any) -> list[str]:
    if value is None or not str(value).strip():
        return []
    parts = [part.strip() for part in re.split(r"[,，]", str(value)) if part.strip()]
    return [re.sub(r"\s*[（(]\d+[）)]\s*$", "", part).strip() for part in parts]


def clean_teacher(value: Any) -> tuple[str | None, bool]:
    names = clean_teacher_names(value)
    if not names:
        return None, False
    return ",".join(names), len(names) > 1


def parse_source_grade(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in CHINESE_GRADES:
        return CHINESE_GRADES[text]
    match = re.fullmatch(r"([1-6])(?:年级)?", text)
    return int(match.group(1)) if match else None


def range_value(mapping: dict[str, Any], grade: int) -> Any:
    for grade_range, value in mapping.items():
        match = re.fullmatch(r"([1-6])-([1-6])", grade_range)
        if match and int(match.group(1)) <= grade <= int(match.group(2)):
            return value
    return None


def workbook_security_issues(path: Path, *, allow_formulas: bool = False) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if path.suffix.lower() != ".xlsx":
        add_issue(issues, BLOCKING, "INVALID_FILE_TYPE", "只接受 .xlsx 文件")
        return issues
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            has_external_target = any(name.startswith("xl/externalLinks/") for name in names)
            if not has_external_target:
                has_external_target = any(
                    re.search(rb"TargetMode\s*=\s*[\"']External[\"']", archive.read(name), re.IGNORECASE)
                    for name in names
                    if name.endswith(".rels")
                )
            if has_external_target:
                add_issue(issues, BLOCKING, "EXTERNAL_LINK", "工作簿包含外部链接")
    except zipfile.BadZipFile:
        add_issue(issues, BLOCKING, "INVALID_XLSX", "文件不是有效的 xlsx 工作簿")
        return issues
    if allow_formulas:
        return issues
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f":
                        add_issue(issues, BLOCKING, "EXCEL_FORMULA", "输入不得包含公式", season=sheet.title, source_ref=cell.coordinate)
                    elif isinstance(value, str) and value.strip().startswith(FORMULA_PREFIXES):
                        add_issue(issues, BLOCKING, "FORMULA_INJECTION", "文本疑似公式注入", season=sheet.title, source_ref=cell.coordinate)
    finally:
        workbook.close()
    return issues


def manual_template_security_issues(path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    issues = workbook_security_issues(path, allow_formulas=True)
    if any(issue["severity"] == BLOCKING for issue in issues):
        return issues
    schedule_sheets = {f"{spec['label']}课表" for spec in config["source"]["seasons"].values()}
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            schedule_season = next(
                (
                    season
                    for season, spec in config["source"]["seasons"].items()
                    if sheet.title == f"{spec['label']}课表"
                ),
                None,
            )
            valid_time_columns = [
                cell.column
                for cell in sheet[1][3:]
                if schedule_season and valid_manual_time_key(str(cell.value or "").strip(), schedule_season, config)
            ]
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f":
                        formula = str(value or "")
                        same_row = str(cell.row)
                        expected_schedule_formula = (
                            f"=COUNTA(D{same_row}:{get_column_letter(max(valid_time_columns))}{same_row})"
                            if valid_time_columns
                            else ""
                        )
                        allowed = (
                            sheet.title in schedule_sheets
                            and cell.column == 3
                            and formula.upper() == expected_schedule_formula.upper()
                        ) or (
                            sheet.title == "容量明细"
                            and cell.column == 11
                            and formula.upper() == f"=I{same_row}-J{same_row}"
                        ) or (
                            sheet.title == "容量汇总"
                            and cell.column == 9
                            and formula.upper() == f"=G{same_row}-H{same_row}"
                        ) or (
                            sheet.title == "容量总览"
                            and cell.column == 7
                            and formula.upper() == f"=E{same_row}-F{same_row}"
                        )
                        if not allowed:
                            add_issue(
                                issues,
                                BLOCKING,
                                "UNEXPECTED_MANUAL_FORMULA",
                                "人工模板包含非本技能生成的公式",
                                season=sheet.title,
                                source_ref=cell.coordinate,
                            )
                    elif isinstance(value, str) and value.strip().startswith(FORMULA_PREFIXES):
                        add_issue(
                            issues,
                            BLOCKING,
                            "FORMULA_INJECTION",
                            "人工模板文本疑似公式注入",
                            season=sheet.title,
                            source_ref=cell.coordinate,
                        )
    finally:
        workbook.close()
    return issues


def resolve_campus_scope(
    config: dict[str, Any],
    campus: str | list[str] | tuple[str, ...] | None,
    campus_group: str | None,
) -> tuple[str, list[str]]:
    if campus_group and campus:
        raise ValueError("--campus 与 --campus-group 不能同时使用")
    if campus_group:
        groups = config["business"].get("campus_groups", {})
        if campus_group not in groups:
            raise ValueError(f"配置中不存在校区管理单元“{campus_group}”")
        return campus_group, list(groups[campus_group])
    if isinstance(campus, str):
        campuses = [campus]
    else:
        campuses = list(campus or [])
    campuses = list(dict.fromkeys(item.strip() for item in campuses if item and item.strip()))
    if not campuses:
        raise ValueError("必须至少提供一个 --campus，或提供 --campus-group")
    return "、".join(campuses), campuses


def read_source(path: Path, config: dict[str, Any], campuses: list[str], issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    records: list[dict[str, Any]] = []
    campus_set = set(campuses)
    try:
        for season, spec in config["source"]["seasons"].items():
            sheet_name = spec["sheet"]
            label = spec["label"]
            if sheet_name not in workbook.sheetnames:
                add_issue(issues, BLOCKING, "MISSING_SHEET", f"缺少工作表 {sheet_name}", season=label)
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if header_row is None:
                add_issue(issues, BLOCKING, "EMPTY_SHEET", "工作表为空", season=label)
                continue
            headers = [str(value).strip() if value is not None else "" for value in header_row]
            required_columns = spec["columns"]
            missing = [name for name in required_columns.values() if name not in headers]
            for column in missing:
                add_issue(issues, BLOCKING, "MISSING_COLUMN", f"缺少列 {column}", season=label)
            if missing:
                continue
            indexes = {key: headers.index(name) for key, name in required_columns.items()}
            optional_indexes = {
                key: headers.index(name)
                for key, name in spec.get("optional_columns", {}).items()
                if name in headers
            }
            for source_row, row in enumerate(rows, start=2):
                if all(value is None or str(value).strip() == "" for value in row):
                    continue
                row_campus = row[indexes["campus"]] if indexes["campus"] < len(row) else None
                normalized_campus = str(row_campus or "").strip()
                if normalized_campus not in campus_set:
                    continue
                record = {
                    key: row[index] if index < len(row) else None
                    for key, index in indexes.items()
                }
                record.update(
                    {
                        key: row[index] if index < len(row) else None
                        for key, index in optional_indexes.items()
                    }
                )
                record.update(
                    {
                        "season": season,
                        "season_label": label,
                        "mode": spec["mode"],
                        "source_campus": normalized_campus,
                        "source_ref": f"{sheet_name}!{source_row}",
                    }
                )
                records.append(record)
    finally:
        workbook.close()
    return records


def parse_time_slot(value: Any, config: dict[str, Any]) -> str | None:
    match = re.search(r"(?<!\d)(\d{1,2})(?:[:：]\d{2}|点)", str(value or ""))
    if not match:
        return None
    hour = int(match.group(1))
    for spec in config["business"]["time_slots"]:
        if spec["minimum_hour"] <= hour <= spec["maximum_hour"]:
            return spec["label"]
    return None


def parse_round(value: Any, start_date: Any, config: dict[str, Any]) -> str | None:
    text = str(value or "").strip()
    round_aliases = {"1": "一轮", "一": "一轮", "一轮": "一轮", "第1轮": "一轮", "2": "二轮", "二": "二轮", "二轮": "二轮", "第2轮": "二轮", "3": "三轮", "三": "三轮", "三轮": "三轮", "第3轮": "三轮"}
    if text in round_aliases:
        return round_aliases[text]
    parsed = parse_date(start_date)
    if parsed is None:
        return None
    point = parsed.strftime("%m-%d")
    for label, (start, end) in config["business"].get("summer_round_dates", {}).items():
        if start <= point <= end:
            return label
    return None


def parse_schedule_date(value: Any) -> date | None:
    match = re.search(r"(?<!\d)(20\d{2})[./-](\d{1,2})[./-](\d{1,2})(?!\d)", str(value or ""))
    if not match:
        return None
    try:
        return date(*(int(part) for part in match.groups()))
    except ValueError:
        return None


def parse_round_time(
    round_value: Any,
    start_date: Any,
    schedule_text: Any,
    config: dict[str, Any],
) -> tuple[str | None, str | None, str]:
    text = str(schedule_text or "")
    markers = config["business"].get("summer_intensive_markers", [])
    segments = [segment.strip() for segment in re.split(r"[;；\n]+", text) if segment.strip()]
    intensive_segments = [segment for segment in segments if any(marker in segment for marker in markers)]
    candidates: set[tuple[str, str]] = set()
    for segment in intensive_segments:
        slot = parse_time_slot(segment, config)
        segment_date = parse_schedule_date(segment)
        period = parse_round(round_value, segment_date or start_date, config)
        if period and slot:
            candidates.add((period, slot))
    if len(candidates) == 1:
        period, slot = next(iter(candidates))
        basis = "intensive_segment" if len(intensive_segments) < len(segments) else "intensive_only"
        return period, slot, basis
    if len(candidates) > 1:
        return None, None, "ambiguous_intensive_segments"
    return parse_round(round_value, start_date, config), parse_time_slot(schedule_text, config), "default"


def parse_weekday(value: Any) -> str | None:
    match = re.search(r"(?:每周|周|星期)([一二三四五六日天1234567])", str(value or ""))
    return WEEKDAY_MAP.get(match.group(1)) if match else None


def source_name_grade(value: Any) -> int | None:
    text = str(value or "")
    for label, grade in CHINESE_GRADES.items():
        if label in text:
            return grade
    match = re.search(r"([1-6])年级", text)
    return int(match.group(1)) if match else None


def source_name_difficulty(value: Any) -> str | None:
    text = str(value or "").upper()
    for code in ("A", "B", "C", "S"):
        if re.search(rf"(?:^|[^A-Z]){code}\s*(?:班型|班|[（(])", text) or re.search(rf"[（(]{code}[）)]", text):
            return code
    return None


def parse_official_code(code: str, config: dict[str, Any]) -> tuple[str | None, int | None, str | None, str | None]:
    prefix = code[:2].upper()
    subject = config["business"]["subject_by_prefix"].get(prefix)
    if subject is None:
        return None, None, None, "不支持的科目编码"
    if len(code) < 3 or code[2] not in "123456":
        return subject, None, None, "班号无法解析为小学一至六年级"
    grade = int(code[2])
    low_grade_nl = config["business"]["low_grade_nl"]
    if prefix == "NL" and grade in low_grade_nl["grades"]:
        difficulty = low_grade_nl["advanced_label"] if code.startswith(f"NL{grade}{low_grade_nl['advanced_suffix']}") else low_grade_nl["base_label"]
        return subject, grade, difficulty, None
    difficulty = code[3].upper() if len(code) > 3 else ""
    if difficulty not in config["business"]["difficulty_codes"]:
        return subject, grade, None, "班号缺少可识别的难度编码"
    return subject, grade, difficulty, None


def normalize_records(raw: list[dict[str, Any]], config: dict[str, Any], issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    business = config["business"]
    normalized: list[dict[str, Any]] = []
    capacity_conflicts: Counter[tuple[str, str, int, int, int]] = Counter()
    code_counts = Counter(str(record.get("class_code") or "").strip() for record in raw if str(record.get("class_code") or "").strip())
    for code, count in code_counts.items():
        if count > 1:
            add_issue(issues, BLOCKING, "DUPLICATE_CLASS_CODE", f"班级编码重复 {count} 次", class_code=code)

    for record in raw:
        season = record["season"]
        label = record["season_label"]
        source_ref = record["source_ref"]
        code = str(record.get("class_code") or "").strip().upper()
        name = str(record.get("class_name") or "").strip()
        if not code:
            add_issue(issues, BLOCKING, "MISSING_CLASS_CODE", "班级编码为空", season=label, source_ref=source_ref)
            continue
        remarks = str(record.get("remarks") or "")
        matched_keywords = [keyword for keyword in business["excluded_name_keywords"] if keyword in name or keyword in remarks]
        if matched_keywords:
            add_issue(issues, EXCLUDED, "EXCLUDED_CLASS_NAME", f"班级名称或备注命中排除词：{'、'.join(matched_keywords)}", season=label, source_ref=source_ref, class_code=code)
            continue
        status = str(record.get("open_status") or "").strip()
        if status in business["closed_status_values"]:
            add_issue(issues, EXCLUDED, "CLOSED_CLASS", f"班级状态“{status}”不进入正式课表", season=label, source_ref=source_ref, class_code=code)
            continue
        source_grade_text = str(record.get("source_grade") or "").strip()
        if source_grade_text in business["excluded_grade_labels"]:
            add_issue(issues, EXCLUDED, "EXCLUDED_GRADE", "S2/S3 不在小学统计范围", season=label, source_ref=source_ref, class_code=code)
            continue
        subject, grade, difficulty, code_error = parse_official_code(code, config)
        if subject is None:
            add_issue(issues, EXCLUDED, "UNSUPPORTED_SUBJECT", code_error or "不支持的科目", season=label, source_ref=source_ref, class_code=code)
            continue
        if grade is None or difficulty is None:
            add_issue(issues, PENDING, "UNRECOGNIZED_CLASS_CODE", code_error or "班号无法解析", season=label, source_ref=source_ref, class_code=code, action="核对班号规则")
            continue
        source_subject = str(record.get("subject") or "").strip()
        aliases = business["subject_aliases"].get(subject, [])
        if source_subject and not any(alias.lower() in source_subject.lower() for alias in aliases):
            add_issue(issues, WARNING, "SUBJECT_CONFLICT", f"源科目“{source_subject}”与班号科目“{subject}”不一致，采用班号", season=label, source_ref=source_ref, class_code=code)
        explicit_grade = parse_source_grade(record.get("source_grade"))
        if explicit_grade is not None and explicit_grade != grade:
            add_issue(issues, WARNING, "GRADE_CONFLICT", f"源年级与班号年级不一致，采用班号 {grade} 年级", season=label, source_ref=source_ref, class_code=code)
        name_grade = source_name_grade(name)
        if name_grade is not None and name_grade != grade:
            add_issue(issues, WARNING, "NAME_GRADE_CONFLICT", f"班级名年级与班号年级不一致，采用班号 {grade} 年级", season=label, source_ref=source_ref, class_code=code)
        name_difficulty = source_name_difficulty(name)
        low_grade_nl = business["low_grade_nl"]
        expected_name_difficulty = low_grade_nl["advanced_suffix"] if difficulty == low_grade_nl["advanced_label"] else difficulty
        if name_difficulty and name_difficulty != expected_name_difficulty:
            add_issue(issues, WARNING, "DIFFICULTY_CONFLICT", f"班级名称难度 {name_difficulty} 与班号难度 {difficulty} 不一致，采用班号", season=label, source_ref=source_ref, class_code=code)

        current = parse_integer(record.get("current_students"))
        if current is None or current < 0:
            add_issue(issues, BLOCKING, "INVALID_CURRENT_STUDENTS", "当前人数必须为非负整数", season=label, source_ref=source_ref, class_code=code)
            continue
        sessions = parse_integer(record.get("class_sessions"))
        expected_sessions = range_value(business["expected_sessions"][season], grade)
        if sessions is None or sessions != expected_sessions:
            add_issue(issues, PENDING, "NONSTANDARD_SESSIONS", f"课次为 {record.get('class_sessions')}，标准课次为 {expected_sessions}", season=label, source_ref=source_ref, class_code=code, action="确认是否属于正常长期班")
            continue
        teacher, multiple = clean_teacher(record.get("teacher"))
        if teacher is None:
            add_issue(issues, PENDING, "MISSING_TEACHER", "班级缺少授课教师", season=label, source_ref=source_ref, class_code=code, action="确认教师后再入表")
            continue
        if multiple:
            teacher_names = clean_teacher_names(record.get("teacher"))
            primary_names = clean_teacher_names(record.get("primary_teacher"))
            if len(primary_names) != 1:
                add_issue(
                    issues,
                    PENDING,
                    "MULTIPLE_TEACHERS_MISSING_PRIMARY",
                    "班级包含多名授课教师，但主带课老师为空或不唯一",
                    season=label,
                    source_ref=source_ref,
                    class_code=code,
                    teacher=teacher,
                    action="在教务系统中确认唯一主带课老师",
                )
                continue
            primary_teacher = primary_names[0]
            if primary_teacher not in teacher_names:
                add_issue(
                    issues,
                    PENDING,
                    "PRIMARY_TEACHER_MISMATCH",
                    "主带课老师不在授课教师列表中",
                    season=label,
                    source_ref=source_ref,
                    class_code=code,
                    teacher=primary_teacher,
                    action=f"核对授课教师：{teacher}",
                )
                continue
            teacher = primary_teacher
            add_issue(
                issues,
                INFO,
                "PRIMARY_TEACHER_SELECTED",
                "班级包含多名授课教师，已按主带课老师入表",
                season=label,
                source_ref=source_ref,
                class_code=code,
                teacher=teacher,
                action="无需人工拆分",
            )
        parsing_basis = "default"
        if record["mode"] == "round":
            period, slot, parsing_basis = parse_round_time(
                record.get("round"),
                record.get("start_date"),
                record.get("schedule_text"),
                config,
            )
        else:
            slot = parse_time_slot(record.get("schedule_text"), config)
            period = parse_weekday(record.get("schedule_text"))
        if parsing_basis == "ambiguous_intensive_segments":
            add_issue(
                issues,
                PENDING,
                "AMBIGUOUS_INTENSIVE_SCHEDULE",
                "存在多个无法归并到同一轮次时段的连续上课片段",
                season=label,
                source_ref=source_ref,
                class_code=code,
                teacher=teacher,
                action="确认应登记的暑假轮次和主时段",
            )
            continue
        if slot is None:
            add_issue(issues, PENDING, "UNRECOGNIZED_TIME", "无法识别上课时段", season=label, source_ref=source_ref, class_code=code, teacher=teacher)
            continue
        if period is None:
            add_issue(issues, PENDING, "UNRECOGNIZED_PERIOD", "无法识别轮次或星期", season=label, source_ref=source_ref, class_code=code, teacher=teacher)
            continue
        if parsing_basis == "intensive_segment":
            add_issue(
                issues,
                WARNING,
                "HYBRID_SCHEDULE_RESOLVED",
                "同时存在周课与连续课，已按连续上课片段识别暑假轮次和时段",
                season=label,
                source_ref=source_ref,
                class_code=code,
                teacher=teacher,
                time_key=f"{period}{slot}",
            )
        time_key = f"{period}{slot}"
        capacity = range_value(business["capacity_rules"][subject], grade)
        if capacity is None:
            add_issue(issues, BLOCKING, "MISSING_CAPACITY_RULE", "缺少科目年级容量规则", season=label, source_ref=source_ref, class_code=code)
            continue
        source_max = parse_integer(record.get("source_max_students"))
        if source_max is not None and source_max != capacity["maximum"]:
            capacity_conflicts[(label, subject, grade, source_max, capacity["maximum"])] += 1
        normalized.append(
            {
                "season": season,
                "season_label": label,
                "mode": record["mode"],
                "subject": subject,
                "grade": grade,
                "difficulty": difficulty,
                "teacher": teacher,
                "teacher_basis": "主带课老师" if multiple else "授课教师",
                "time_key": time_key,
                "class_code": code,
                "current_students": current,
                "minimum": capacity["minimum"],
                "maximum": capacity["maximum"],
                "source_type": "正式班",
                "source_campus": record["source_campus"],
                "source_ref": source_ref,
            }
        )
    for (label, subject, grade, source_max, configured_max), count in sorted(capacity_conflicts.items()):
        add_issue(
            issues,
            WARNING,
            "SOURCE_MAX_CONFLICT",
            f"{subject}{grade}年级有 {count} 个班的源最大人数为 {source_max}，配置最大班容为 {configured_max}，统一采用配置",
            season=label,
            action="如业务班容已调整，请修改版本化配置后重新运行",
            affected_count=count,
        )
    return normalized


def parse_planned_classes(
    path: Path,
    config: dict[str, Any],
    official_records: dict[str, dict[str, Any]],
    source_codes: set[str],
    seen_official_codes: set[tuple[str, str]],
    compared_seasons: set[str],
    template_positions: dict[tuple[str, str], dict[str, str] | None],
    template_policy: str,
    issues: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    planned: list[dict[str, Any]] = []
    pattern = re.compile(config["business"]["planned_class_pattern"])
    try:
        for season, spec in config["source"]["seasons"].items():
            sheet_name = f"{spec['label']}课表"
            if sheet_name not in workbook.sheetnames:
                continue
            compared_seasons.add(season)
            sheet = workbook[sheet_name]
            headers = [str(cell.value or "").strip() for cell in sheet[1]]
            for row in range(2, sheet.max_row + 1):
                row_subject = str(sheet.cell(row, 1).value or "").strip()
                teacher = str(sheet.cell(row, 2).value or "").strip()
                for column in range(4, sheet.max_column + 1):
                    cell = sheet.cell(row, column)
                    value = str(cell.value or "").strip().upper()
                    if value.startswith("新班"):
                        match = pattern.fullmatch(value)
                        source_ref = f"{path.name}:{sheet_name}!{cell.coordinate}"
                        if not match:
                            add_issue(issues, PENDING, "INVALID_PLANNED_CLASS", "计划新班格式不符合约定", season=spec["label"], source_ref=source_ref, class_code=value, teacher=teacher)
                            continue
                        prefix, grade_text, suffix = match.groups()
                        grade = int(grade_text)
                        subject = config["business"]["subject_by_prefix"][prefix]
                        low_grade_nl = config["business"]["low_grade_nl"]
                        if prefix == "NL" and grade in low_grade_nl["grades"]:
                            if suffix == low_grade_nl["planned_base_suffix"]:
                                difficulty = low_grade_nl["base_label"]
                            elif suffix == low_grade_nl["advanced_suffix"]:
                                difficulty = low_grade_nl["advanced_label"]
                            else:
                                add_issue(issues, PENDING, "INVALID_NL_PLANNED_SUFFIX", "NL1/NL2 只接受无后缀或 A 后缀", season=spec["label"], source_ref=source_ref, class_code=value, teacher=teacher)
                                continue
                        else:
                            if suffix not in config["business"]["difficulty_codes"]:
                                add_issue(issues, PENDING, "MISSING_PLANNED_DIFFICULTY", "计划新班缺少有效难度字母", season=spec["label"], source_ref=source_ref, class_code=value, teacher=teacher)
                                continue
                            difficulty = suffix
                        if not teacher:
                            add_issue(issues, PENDING, "PLANNED_CLASS_MISSING_TEACHER", "计划新班所在行缺少教师", season=spec["label"], source_ref=source_ref, class_code=value)
                            continue
                        if row_subject and row_subject != subject:
                            add_issue(issues, WARNING, "PLANNED_SUBJECT_CONFLICT", f"行科目“{row_subject}”与新班编码科目“{subject}”不一致，采用编码", season=spec["label"], source_ref=source_ref, class_code=value, teacher=teacher)
                        time_key = headers[column - 1] if column - 1 < len(headers) else ""
                        if not valid_manual_time_key(time_key, season, config):
                            add_issue(issues, PENDING, "PLANNED_CLASS_INVALID_TIME", "计划新班所在列不是可识别的课表时段", season=spec["label"], source_ref=source_ref, class_code=value, teacher=teacher)
                            continue
                        capacity = range_value(config["business"]["capacity_rules"][subject], grade)
                        if capacity is None:
                            add_issue(issues, BLOCKING, "MISSING_CAPACITY_RULE", "计划新班缺少科目年级容量规则", season=spec["label"], source_ref=source_ref, class_code=value)
                            continue
                        planned.append(
                            {
                                "season": season,
                                "season_label": spec["label"],
                                "mode": spec["mode"],
                                "subject": subject,
                                "grade": grade,
                                "difficulty": difficulty,
                                "teacher": teacher,
                                "time_key": time_key,
                                "class_code": value,
                                "current_students": 0,
                                "minimum": capacity["minimum"],
                                "maximum": capacity["maximum"],
                                "source_type": "计划新班",
                                "source_campus": "",
                                "source_ref": source_ref,
                            }
                        )
                    elif "\n" in value:
                        code_match = re.search(r"\n\s*([A-Z]{2}[A-Z0-9]+)\s*$", value)
                        if not code_match:
                            continue
                        official_code = code_match.group(1)
                        subject, grade, difficulty, _ = parse_official_code(official_code, config)
                        if subject is None or grade is None or difficulty is None:
                            continue
                        seen_official_codes.add((season, official_code))
                        time_key = headers[column - 1]
                        official = official_records.get(official_code)
                        source_ref = f"{path.name}:{sheet_name}!{cell.coordinate}"
                        if official_code not in source_codes:
                            add_issue(issues, PENDING, "STALE_TEMPLATE_CLASS", "人工模板班号不在本次源数据中，疑似取消班或陈旧数据", season=spec["label"], source_ref=source_ref, class_code=official_code, teacher=teacher, time_key=time_key, action="确认删除或改写为计划新班")
                        elif official is None:
                            add_issue(issues, PENDING, "TEMPLATE_CLASS_NOT_ELIGIBLE", "人工模板班号仍在源数据中，但未通过当前统计规则", season=spec["label"], source_ref=source_ref, class_code=official_code, teacher=teacher, time_key=time_key, action="查看待确认或排除记录后决定是否保留")
                        elif official["season"] != season:
                            add_issue(
                                issues,
                                PENDING,
                                "TEMPLATE_SEASON_MISMATCH",
                                "人工模板所在季节与教务导出不一致，季节不能由模板覆盖",
                                season=spec["label"],
                                source_ref=source_ref,
                                class_code=official_code,
                                teacher=teacher,
                                time_key=time_key,
                                action=f"放回 {official['season_label']}课表或更新教务数据",
                            )
                        elif not teacher:
                            add_issue(
                                issues,
                                PENDING,
                                "TEMPLATE_CLASS_MISSING_TEACHER",
                                "人工模板中的正式班所在行缺少教师",
                                season=spec["label"],
                                source_ref=source_ref,
                                class_code=official_code,
                                time_key=time_key,
                                action="补充教师后重新运行",
                            )
                        elif not valid_manual_time_key(time_key, season, config):
                            add_issue(
                                issues,
                                PENDING,
                                "TEMPLATE_CLASS_INVALID_TIME",
                                "人工模板中的正式班所在列不是可识别时段",
                                season=spec["label"],
                                source_ref=source_ref,
                                class_code=official_code,
                                teacher=teacher,
                                time_key=time_key,
                                action="移到标准时段列后重新运行",
                            )
                        else:
                            key = (season, official_code)
                            position = {"teacher": teacher, "time_key": time_key, "source_ref": source_ref}
                            existing_position = template_positions.get(key)
                            position_changed = existing_position is not None and any(
                                existing_position[field] != position[field] for field in ("teacher", "time_key")
                            )
                            if key in template_positions and (existing_position is None or position_changed):
                                template_positions[key] = None
                                add_issue(
                                    issues,
                                    PENDING,
                                    "DUPLICATE_TEMPLATE_CLASS",
                                    "同一正式班在人工模板中出现多个不同位置",
                                    season=spec["label"],
                                    source_ref=source_ref,
                                    class_code=official_code,
                                    teacher=teacher,
                                    time_key=time_key,
                                    action="只保留一个确定位置后重新运行",
                                )
                            elif key not in template_positions:
                                template_positions[key] = position
                            if (
                                template_policy == "rule_first"
                                and (official["teacher"] != teacher or official["time_key"] != time_key)
                            ):
                                add_issue(
                                    issues,
                                    PENDING,
                                    "MANUAL_OFFICIAL_CLASS_MOVE",
                                    "人工模板中的正式班教师或时段与教务导出不一致，规则优先模式下未应用",
                                    season=spec["label"],
                                    source_ref=source_ref,
                                    class_code=official_code,
                                    teacher=teacher,
                                    time_key=time_key,
                                    action=f"如需采用人工位置，改用 template_first；源位置为 {official['teacher']} / {official['time_key']}",
                                )
    finally:
        workbook.close()
    return planned


def apply_template_policy(
    records: list[dict[str, Any]],
    compared_seasons: set[str],
    template_positions: dict[tuple[str, str], dict[str, str] | None],
    template_policy: str,
    issues: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if template_policy not in {"rule_first", "template_first"}:
        raise ValueError("template_policy 必须是 rule_first 或 template_first")
    if template_policy == "rule_first":
        return records
    applied: list[dict[str, Any]] = []
    for record in records:
        if record["season"] not in compared_seasons:
            applied.append(record)
            continue
        key = (record["season"], record["class_code"])
        if key not in template_positions:
            add_issue(
                issues,
                INFO,
                "TEMPLATE_OVERRIDE_SOURCE_REMOVED",
                "正式班未出现在已回传模板中，模板优先模式下不再入表",
                season=record["season_label"],
                source_ref=record["source_ref"],
                class_code=record["class_code"],
                teacher=record["teacher"],
                time_key=record["time_key"],
                action="若为模板漏登，补回班号后重新运行；若要保留数据源，改用 rule_first",
            )
            continue
        position = template_positions[key]
        if position is None:
            continue
        updated = dict(record)
        changed = updated["teacher"] != position["teacher"] or updated["time_key"] != position["time_key"]
        updated["teacher"] = position["teacher"]
        updated["time_key"] = position["time_key"]
        updated["template_source_ref"] = position["source_ref"]
        updated["template_override"] = changed
        applied.append(updated)
        if changed:
            add_issue(
                issues,
                INFO,
                "TEMPLATE_OVERRIDE_APPLIED",
                "已按模板优先开关应用正式班的教师或时段变更",
                season=record["season_label"],
                source_ref=position["source_ref"],
                class_code=record["class_code"],
                teacher=position["teacher"],
                time_key=position["time_key"],
                action=f"源位置为 {record['teacher']} / {record['time_key']}",
            )
    return applied


def resolve_conflicts(records: list[dict[str, Any]], issues: list[dict[str, Any]], config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    policy = config["business"]["conflict_policy"]
    if policy["multiple_classes"] != "unique_max_current" or policy["highest_tie"] != "manual_confirmation":
        raise ValueError("未支持的教师时段冲突策略")
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        groups[(record["season"], record["teacher"], record["time_key"])].append(record)
    selected: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    for (_, teacher, time_key), candidates in sorted(groups.items()):
        if len(candidates) == 1:
            candidate = dict(candidates[0])
            candidate["selected"] = True
            selected.append(candidate)
            continue
        maximum = max(candidate["current_students"] for candidate in candidates)
        winners = [candidate for candidate in candidates if candidate["current_students"] == maximum]
        unique_winner = winners[0] if len(winners) == 1 else None
        severity = WARNING if unique_winner else PENDING
        code = "TEACHER_TIME_CONFLICT_SELECTED" if unique_winner else "TEACHER_TIME_CONFLICT_TIE"
        message = "同一教师同一时段多班，临时选择当前人数最多的班" if unique_winner else "同一教师同一时段多班且最高人数相同，必须人工确认"
        add_issue(issues, severity, code, message, season=candidates[0]["season_label"], teacher=teacher, time_key=time_key, action="调整模板或确认保留班级")
        for candidate in candidates:
            chosen = candidate is unique_winner
            conflicts.append(
                {
                    "season_label": candidate["season_label"],
                    "teacher": teacher,
                    "time_key": time_key,
                    "class_code": candidate["class_code"],
                    "current_students": candidate["current_students"],
                    "selected": "是" if chosen else "否",
                    "reason": message,
                }
            )
            if chosen:
                record = dict(candidate)
                record["selected"] = True
                selected.append(record)
    return selected, conflicts


def status_color(record: dict[str, Any], colors: dict[str, str]) -> str:
    current = record["current_students"]
    if current < record["minimum"]:
        return colors["below_minimum"]
    if current < record["maximum"]:
        return colors["open_not_full"]
    return colors["full_or_oversold"]


def safe_cell(value: Any) -> Any:
    return "'" + value if isinstance(value, str) and value.startswith(FORMULA_PREFIXES) else value


def style_simple_table(
    sheet,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    colors: dict[str, str],
    *,
    formula_columns: set[int] | None = None,
    empty_message: str | None = None,
) -> None:
    last_column = get_column_letter(max(1, len(headers)))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=colors["header"])
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    border = Side(style="thin", color=colors["border"])
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(name="Microsoft YaHei", size=10.5, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=colors["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=border)
    for row_number, values in enumerate(rows, start=4):
        for column, value in enumerate(values, start=1):
            trusted_formula = formula_columns and column in formula_columns and isinstance(value, str) and value.startswith("=")
            cell = sheet.cell(row_number, column, value if trusted_formula else safe_cell(value))
            cell.font = Font(name="Microsoft YaHei", size=10, color=colors["text"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color=colors["border"]))
    if not rows and empty_message:
        sheet.merge_cells(f"A4:{last_column}4")
        cell = sheet["A4"]
        cell.value = empty_message
        cell.font = Font(name="Microsoft YaHei", size=10.5, italic=True, color=colors["text"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=colors["teacher_fill"])
        sheet.row_dimensions[4].height = 26
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A3:{last_column}{max(3, len(rows) + 3)}"
    for column, header in enumerate(headers, start=1):
        contents = [str(header)] + [str(row[column - 1] or "") for row in rows]
        sheet.column_dimensions[get_column_letter(column)].width = min(36, max(10, max(map(len, contents), default=8) + 2))


def schedule_columns(season: str, records: list[dict[str, Any]], config: dict[str, Any]) -> list[str]:
    mode = config["source"]["seasons"][season]["mode"]
    if mode == "round":
        return [f"{round_label}{slot['label']}" for round_label in config["business"]["rounds"] for slot in config["business"]["time_slots"]]
    columns = set(config["business"]["weekly_default_columns"])
    columns.update(record["time_key"] for record in records if record["season"] == season)
    day_order = {value: index for index, value in enumerate(config["business"]["weekday_order"])}
    slot_order = {value["label"]: index for index, value in enumerate(config["business"]["time_slots"])}

    def sort_key(value: str) -> tuple[int, int, str]:
        day = next((item for item in day_order if value.startswith(item)), "")
        slot = value[len(day):]
        return day_order.get(day, 99), slot_order.get(slot, 99), value

    return sorted(columns, key=sort_key)


def valid_manual_time_key(time_key: str, season: str, config: dict[str, Any]) -> bool:
    if config["source"]["seasons"][season]["mode"] == "round":
        return time_key in {
            f"{round_label}{slot['label']}"
            for round_label in config["business"]["rounds"]
            for slot in config["business"]["time_slots"]
        }
    day = next((label for label in config["business"]["weekday_order"] if time_key.startswith(label)), "")
    slot = time_key[len(day):]
    return bool(day) and slot in {item["label"] for item in config["business"]["time_slots"]}


def time_fill(time_key: str, season: str, config: dict[str, Any]) -> str:
    colors = config["colors"]
    if config["source"]["seasons"][season]["mode"] == "round":
        for index, label in enumerate(config["business"]["rounds"]):
            if time_key.startswith(label):
                return colors["round_fills"][index]
    if time_key.startswith("周六"):
        return colors["saturday_fill"]
    if time_key.startswith("周日"):
        return colors["sunday_fill"]
    return colors["weekday_fill"]


def write_schedule_sheet(workbook: Workbook, season: str, records: list[dict[str, Any]], config: dict[str, Any]) -> None:
    spec = config["source"]["seasons"][season]
    colors = config["colors"]
    sheet = workbook.create_sheet(f"{spec['label']}课表")
    times = schedule_columns(season, records, config)
    headers = ["科目", "教师", "班量", *times]
    border = Side(style="thin", color=colors["border"])
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color=colors["header_text"])
        cell.fill = PatternFill("solid", fgColor=colors["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=border, right=border, top=border, bottom=border)
    season_records = [record for record in records if record["season"] == season]
    groups: dict[tuple[str, str], dict[str, dict[str, Any]]] = defaultdict(dict)
    for record in season_records:
        groups[(record["subject"], record["teacher"])][record["time_key"]] = record
    subject_order = {subject: index for index, subject in enumerate(config["business"]["subject_order"])}
    for row_number, ((subject, teacher), occupied) in enumerate(sorted(groups.items(), key=lambda item: (subject_order.get(item[0][0], 99), item[0][1])), start=2):
        sheet.cell(row_number, 1, subject)
        sheet.cell(row_number, 2, teacher)
        first_time = get_column_letter(4)
        last_time = get_column_letter(3 + len(times))
        sheet.cell(row_number, 3, f"=COUNTA({first_time}{row_number}:{last_time}{row_number})")
        for column, time_key in enumerate(times, start=4):
            cell = sheet.cell(row_number, column)
            record = occupied.get(time_key)
            if record:
                cell.value = record["class_code"] if record["source_type"] == "计划新班" else f"{record['current_students']}人\n{record['class_code']}"
                cell.font = Font(name="Microsoft YaHei", size=10.5, bold=True, color=status_color(record, colors))
            else:
                cell.font = Font(name="Microsoft YaHei", size=10, color=colors["text"])
            cell.fill = PatternFill("solid", fgColor=time_fill(time_key, season, config))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=border, right=border, top=border, bottom=border)
        for column in (1, 2, 3):
            cell = sheet.cell(row_number, column)
            fill = colors["subject_fill"] if column == 1 else colors["teacher_fill"]
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Microsoft YaHei", size=10.5, bold=column == 1, color=colors["text"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=border, right=border, top=border, bottom=border)
        sheet.row_dimensions[row_number].height = 38
    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 8
    for column in range(4, 4 + len(times)):
        sheet.column_dimensions[get_column_letter(column)].width = 15
    sheet.freeze_panes = "D2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:{get_column_letter(3 + len(times))}{max(1, sheet.max_row)}"


def flatten_config(value: Any, prefix: str = "") -> list[list[str]]:
    rows: list[list[str]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            rows.extend(flatten_config(value[key], f"{prefix}.{key}" if prefix else key))
    elif isinstance(value, list):
        rows.append([prefix, json.dumps(value, ensure_ascii=False)])
    else:
        rows.append([prefix, str(value)])
    return rows


def write_result_workbook(
    destination: Path,
    *,
    status: str,
    data_as_of: str,
    campus_label: str,
    campuses: list[str],
    run_id: str,
    input_hash: str,
    config: dict[str, Any],
    issues: list[dict[str, Any]],
    selected: list[dict[str, Any]],
    conflicts: list[dict[str, Any]],
    source_record_count: int,
    manual_template_count: int,
    template_policy: str,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    colors = config["colors"]
    summary_rows = [
        ["运行状态", status],
        ["数据日期", data_as_of],
        ["管理口径", campus_label],
        ["纳入源校区", "、".join(campuses)],
        ["运行ID", run_id],
        ["输入哈希", input_hash],
        ["配置版本", config["schema_version"]],
        ["输出版本", config["output_version"]],
        ["规则来源", "初始规则+本地覆盖" if config["_rules"]["mode"] == "local_override" else "初始规则"],
        ["模板优先级", template_policy],
        ["阻断问题", sum(issue["severity"] == BLOCKING for issue in issues)],
        ["待确认问题", sum(issue["severity"] == PENDING for issue in issues)],
        ["警告", sum(issue["severity"] == WARNING for issue in issues)],
        ["信息", sum(issue["severity"] == INFO for issue in issues)],
        ["排除记录", sum(issue["severity"] == EXCLUDED for issue in issues)],
        ["源记录", source_record_count],
        ["源班容差异记录", sum(issue["affected_count"] for issue in issues if issue["code"] == "SOURCE_MAX_CONFLICT")],
        ["人工模板", manual_template_count],
        ["数据源新增", sum(issue["code"] == "SOURCE_ONLY_CLASS" for issue in issues)],
        ["模板遗留", sum(issue["code"] == "STALE_TEMPLATE_CLASS" for issue in issues)],
        ["教师/时间变更", sum(issue["code"] == "MANUAL_OFFICIAL_CLASS_MOVE" for issue in issues)],
        ["模板已应用变更", sum(issue["code"] == "TEMPLATE_OVERRIDE_APPLIED" for issue in issues)],
        ["模板已移除源班", sum(issue["code"] == "TEMPLATE_OVERRIDE_SOURCE_REMOVED" for issue in issues)],
        ["有效班级", len(selected)],
        ["人工确认", "必须"],
        ["状态色", f"未开班 #{colors['below_minimum']} / 已开未满 #{colors['open_not_full']} / 满班超售 #{colors['full_or_oversold']}"],
    ]
    summary_sheet = workbook.create_sheet("运行摘要")
    style_simple_table(summary_sheet, "教师课表规整运行摘要", ["指标", "值"], summary_rows, colors)
    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 60
    issue_headers = ["级别", "代码", "季节", "源位置", "班级编码", "教师", "时间", "影响记录", "说明", "建议动作"]
    issue_rows = [[issue[key] for key in ("severity", "code", "season", "source_ref", "class_code", "teacher", "time_key", "affected_count", "message", "action")] for issue in issues]
    validation_rows = [row for row in issue_rows if row[0] != EXCLUDED]
    style_simple_table(
        workbook.create_sheet("校验问题"),
        "数据校验与人工确认问题",
        issue_headers,
        validation_rows,
        colors,
        empty_message="本次未发现校验问题",
    )

    blocked = status == "blocked"
    if not blocked:
        excluded_rows = [row for row in issue_rows if row[0] == EXCLUDED]
        style_simple_table(
            workbook.create_sheet("排除记录"),
            "按规则排除、不进入正式课表的记录",
            issue_headers,
            excluded_rows,
            colors,
            empty_message="本次没有排除记录",
        )
        for season in config["source"]["seasons"]:
            write_schedule_sheet(workbook, season, selected, config)

        details = sorted(selected, key=lambda row: (row["season_label"], row["subject"], row["grade"], str(row["difficulty"]), row["time_key"], row["teacher"], row["class_code"]))
        detail_headers = ["季节", "科目", "年级", "难度/班型", "时间", "班级来源", "教师", "班级编码", "总容量", "当前数量", "差额"]
        detail_rows = []
        for index, record in enumerate(details, start=4):
            detail_rows.append([record["season_label"], record["subject"], record["grade"], record["difficulty"], record["time_key"], record["source_type"], record["teacher"], record["class_code"], record["maximum"], record["current_students"], f"=I{index}-J{index}"])
        overview_groups: dict[tuple[Any, ...], list[int]] = defaultdict(lambda: [0, 0])
        for record in details:
            key = (record["season_label"], record["subject"], record["grade"], record["difficulty"])
            overview_groups[key][0] += record["maximum"]
            overview_groups[key][1] += record["current_students"]
        overview_headers = ["季节", "科目", "年级", "难度/班型", "总容量", "当前数量", "差额"]
        overview_rows = []
        for index, (key, values) in enumerate(sorted(overview_groups.items()), start=4):
            overview_rows.append([*key, values[0], values[1], f"=E{index}-F{index}"])
        style_simple_table(workbook.create_sheet("容量总览"), "按季节、科目、年级和难度汇总", overview_headers, overview_rows, colors, formula_columns={7})
        style_simple_table(workbook.create_sheet("容量明细"), "班级容量明细", detail_headers, detail_rows, colors, formula_columns={11})

        groups: dict[tuple[Any, ...], list[int]] = defaultdict(lambda: [0, 0])
        for record in details:
            key = (record["season_label"], record["subject"], record["grade"], record["difficulty"], record["time_key"], record["source_type"])
            groups[key][0] += record["maximum"]
            groups[key][1] += record["current_students"]
        summary_headers = ["季节", "科目", "年级", "难度/班型", "时间", "班级来源", "总容量", "当前数量", "差额"]
        capacity_rows = []
        for index, (key, values) in enumerate(sorted(groups.items()), start=4):
            capacity_rows.append([*key, values[0], values[1], f"=G{index}-H{index}"])
        style_simple_table(workbook.create_sheet("容量汇总"), "按时间和班级来源下钻的容量汇总", summary_headers, capacity_rows, colors, formula_columns={9})

        conflict_headers = ["季节", "教师", "时间", "班级编码", "当前数量", "临时入表", "原因"]
        conflict_rows = [[row[key] for key in ("season_label", "teacher", "time_key", "class_code", "current_students", "selected", "reason")] for row in conflicts]
        style_simple_table(workbook.create_sheet("排课冲突"), "教师时段排课冲突", conflict_headers, conflict_rows, colors, empty_message="本次未发现教师时段冲突")
        pending_rows = [row for row in issue_rows if row[0] == PENDING]
        style_simple_table(workbook.create_sheet("待确认"), "需要使用者确认的记录", issue_headers, pending_rows, colors, empty_message="本次没有待确认记录")
        comparison_codes = {
            "SOURCE_ONLY_CLASS",
            "STALE_TEMPLATE_CLASS",
            "TEMPLATE_CLASS_NOT_ELIGIBLE",
            "MANUAL_OFFICIAL_CLASS_MOVE",
            "TEMPLATE_OVERRIDE_APPLIED",
            "TEMPLATE_OVERRIDE_SOURCE_REMOVED",
            "DUPLICATE_TEMPLATE_CLASS",
            "TEMPLATE_SEASON_MISMATCH",
        }
        comparison_rows = [row for row in issue_rows if row[1] in comparison_codes]
        comparison_message = "未提供人工模板，未执行三向对比" if manual_template_count == 0 else "本次人工模板与数据源一致"
        style_simple_table(workbook.create_sheet("模板对比"), "数据源与人工模板三向对比", issue_headers, comparison_rows, colors, empty_message=comparison_message)
        style_simple_table(workbook.create_sheet("假设与版本"), "配置、假设与版本", ["路径", "值"], flatten_config(config), colors)

    workbook.properties.creator = "xdf-normalize-teacher-schedule"
    workbook.properties.title = "教师课表占档与容量差额"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()


def run_schedule(
    input_path: Path,
    config_path: Path,
    output_dir: Path,
    *,
    data_as_of: str,
    campus: str | list[str] | tuple[str, ...] | None = None,
    campus_group: str | None = None,
    manual_template: Path | None = None,
    manual_templates: list[Path] | tuple[Path, ...] | None = None,
    template_policy: str = "rule_first",
    rules_overlay: Path | None = None,
) -> dict[str, Any]:
    if template_policy not in {"rule_first", "template_first"}:
        raise ValueError("template_policy 必须是 rule_first 或 template_first")
    config = load_config(config_path, rules_overlay)
    campus_label, campuses = resolve_campus_scope(config, campus, campus_group)
    template_paths = []
    if manual_template:
        template_paths.append(manual_template)
    template_paths.extend(manual_templates or [])
    template_paths = list(dict.fromkeys(Path(path) for path in template_paths))
    parsed_date = parse_date(data_as_of)
    if parsed_date is None or parsed_date.isoformat() != data_as_of:
        raise ValueError("data_as_of 必须为 YYYY-MM-DD")
    issues = workbook_security_issues(input_path)
    raw: list[dict[str, Any]] = []
    if not any(issue["severity"] == BLOCKING for issue in issues):
        raw = read_source(input_path, config, campuses, issues)
        if not raw and not any(issue["severity"] == BLOCKING for issue in issues):
            add_issue(issues, BLOCKING, "NO_CAMPUS_RECORDS", f"未找到管理口径“{campus_label}”包含的校区记录", action="核对完整校区名称、管理单元配置和源文件")
    normalized: list[dict[str, Any]] = []
    if not any(issue["severity"] == BLOCKING for issue in issues):
        normalized = normalize_records(raw, config, issues)
        if not normalized and not any(issue["severity"] in {BLOCKING, PENDING, WARNING} for issue in issues):
            add_issue(issues, PENDING, "NO_ELIGIBLE_CLASSES", "没有符合当前小学素养规则的班级", action="核对校区、季节和统计范围")
    if template_paths:
        for path in template_paths:
            issues.extend(manual_template_security_issues(path, config))
        if not any(issue["severity"] == BLOCKING for issue in issues):
            official_records = {record["class_code"]: record for record in normalized}
            source_codes = {str(record.get("class_code") or "").strip().upper() for record in raw}
            seen_official_codes: set[tuple[str, str]] = set()
            compared_seasons: set[str] = set()
            template_positions: dict[tuple[str, str], dict[str, str] | None] = {}
            planned: list[dict[str, Any]] = []
            for path in template_paths:
                planned.extend(
                    parse_planned_classes(
                        path,
                        config,
                        official_records,
                        source_codes,
                        seen_official_codes,
                        compared_seasons,
                        template_positions,
                        template_policy,
                        issues,
                    )
                )
            if template_policy == "rule_first":
                for record in official_records.values():
                    key = (record["season"], record["class_code"])
                    if record["season"] in compared_seasons and key not in seen_official_codes:
                        add_issue(
                            issues,
                            PENDING,
                            "SOURCE_ONLY_CLASS",
                            "班级存在于本次数据源，但未出现在对应季节的人工模板中",
                            season=record["season_label"],
                            source_ref=record["source_ref"],
                            class_code=record["class_code"],
                            teacher=record["teacher"],
                            time_key=record["time_key"],
                            action="确认是否为新开班，或人工模板是否漏登",
                        )
            normalized = apply_template_policy(
                normalized,
                compared_seasons,
                template_positions,
                template_policy,
                issues,
            )
            normalized.extend(planned)
    selected: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    if not any(issue["severity"] == BLOCKING for issue in issues):
        selected, conflicts = resolve_conflicts(normalized, issues, config)

    blocked = any(issue["severity"] == BLOCKING for issue in issues)
    needs_confirmation = any(issue["severity"] in {PENDING, WARNING} for issue in issues)
    status = "blocked" if blocked else ("needs_confirmation" if needs_confirmation else "ready")
    digest = hashlib.sha256()
    digest.update(input_path.read_bytes())
    digest.update(config_path.read_bytes())
    if rules_overlay is not None:
        digest.update(rules_overlay.read_bytes())
    for path in template_paths:
        digest.update(path.read_bytes())
    digest.update(campus_label.encode("utf-8"))
    digest.update("\0".join(campuses).encode("utf-8"))
    digest.update(data_as_of.encode("utf-8"))
    digest.update(template_policy.encode("utf-8"))
    input_hash = digest.hexdigest()
    run_id = input_hash[:16]
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "teacher-schedule-result.xlsx"
    review_path = output_dir / "schedule-review.json"
    write_result_workbook(
        workbook_path,
        status=status,
        data_as_of=data_as_of,
        campus_label=campus_label,
        campuses=campuses,
        run_id=run_id,
        input_hash=input_hash,
        config=config,
        issues=issues,
        selected=selected,
        conflicts=conflicts,
        source_record_count=len(raw),
        manual_template_count=len(template_paths),
        template_policy=template_policy,
    )
    review = {
        "schema_version": "1.2.0",
        "status": status,
        "run_id": run_id,
        "data_as_of": data_as_of,
        "campus": campus_label,
        "campuses": campuses,
        "config_version": config["schema_version"],
        "output_version": config["output_version"],
        "input_hash": input_hash,
        "rules_mode": config["_rules"]["mode"],
        "base_config_sha256": config["_rules"]["base_config_sha256"],
        "rule_override_count": config["_rules"]["override_count"],
        "template_policy": template_policy,
        "requires_human_confirmation": True,
        "auto_apply": False,
        "template_changes_applied": template_policy == "template_first",
        "human_followup_prompt": "请确认教务导出后是否发生关班、新开班、更换教师或移动时段；可回传人工修改的课表并选择 rule_first 或 template_first。自然语言规则变更只写入本地覆盖，可随时恢复初始规则。",
        "selected_class_count": len(selected),
        "source_record_count": len(raw),
        "excluded_record_count": sum(issue["severity"] == EXCLUDED for issue in issues),
        "manual_template_count": len(template_paths),
        "manual_comparison_counts": {
            "source_only": sum(issue["code"] == "SOURCE_ONLY_CLASS" for issue in issues),
            "template_only": sum(issue["code"] == "STALE_TEMPLATE_CLASS" for issue in issues),
            "source_present_but_ineligible": sum(issue["code"] == "TEMPLATE_CLASS_NOT_ELIGIBLE" for issue in issues),
            "teacher_or_time_changed": sum(issue["code"] == "MANUAL_OFFICIAL_CLASS_MOVE" for issue in issues),
            "template_changes_applied": sum(issue["code"] == "TEMPLATE_OVERRIDE_APPLIED" for issue in issues),
            "source_removed_by_template": sum(issue["code"] == "TEMPLATE_OVERRIDE_SOURCE_REMOVED" for issue in issues),
        },
        "issues": issues,
        "conflicts": conflicts,
        "attachments": [workbook_path.name],
    }
    review_path.write_text(json.dumps(review, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {
        "status": status,
        "run_id": run_id,
        "campus": campus_label,
        "campuses": campuses,
        "issues": issues,
        "selected": selected,
        "conflicts": conflicts,
        "outputs": {"workbook": str(workbook_path), "review": str(review_path)},
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="小学素养教师课表占档规整")
    parser.add_argument("--input", type=Path, required=True, help="教务系统班级明细 xlsx")
    parser.add_argument("--config", type=Path, required=True, help="版本化 JSON 配置")
    campus_scope = parser.add_mutually_exclusive_group(required=True)
    campus_scope.add_argument("--campus", action="append", help="源表中的完整校区名称；可重复提供")
    campus_scope.add_argument("--campus-group", help="配置中的校区管理单元名称")
    parser.add_argument("--data-as-of", required=True, help="教务系统导出日期 YYYY-MM-DD")
    parser.add_argument("--output-dir", type=Path, required=True, help="输出目录")
    parser.add_argument("--manual-template", type=Path, action="append", help="可选：人工修改后的本技能课表；暑假、秋季文件可分别重复提供")
    parser.add_argument("--template-policy", choices=("rule_first", "template_first"), default="rule_first", help="正式班冲突时采用规则或回传模板优先")
    parser.add_argument("--rules-overlay", type=Path, help="可选：本地规则覆盖 JSON；初始配置不会被修改")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = run_schedule(
            args.input,
            args.config,
            args.output_dir,
            campus=args.campus,
            campus_group=args.campus_group,
            data_as_of=args.data_as_of,
            manual_templates=args.manual_template,
            template_policy=args.template_policy,
            rules_overlay=args.rules_overlay,
        )
    except Exception as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps({"status": result["status"], "run_id": result["run_id"], "outputs": result["outputs"]}, ensure_ascii=False))
    return 2 if result["status"] == "blocked" else 0


if __name__ == "__main__":
    raise SystemExit(main())
