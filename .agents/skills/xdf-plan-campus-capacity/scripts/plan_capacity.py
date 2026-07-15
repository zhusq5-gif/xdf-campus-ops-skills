#!/usr/bin/env python3
"""Deterministic campus teacher and classroom capacity planner."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
import zipfile
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLOCKING = "阻断"
WARNING = "警告"
FORMULA_PREFIXES = ("=", "+", "-", "@")
PHONE_RE = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
ID_RE = re.compile(r"(?<!\d)\d{17}[0-9Xx](?!\d)")


def add_issue(
    issues: list[dict[str, Any]],
    code: str,
    message: str,
    *,
    table: str = "",
    row: int | str = "",
    field: str = "",
    severity: str = BLOCKING,
) -> None:
    issues.append(
        {
            "severity": severity,
            "code": code,
            "table": table,
            "row": row,
            "field": field,
            "message": message,
        }
    )


def month_value(value: Any) -> str | None:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    if value is None:
        return None
    text = str(value).strip()
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})(?:[-/]\d{1,2})?", text)
    if not match:
        return None
    year, month = int(match.group(1)), int(match.group(2))
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


def date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def add_month(month: str, offset: int) -> str:
    year, number = (int(part) for part in month.split("-"))
    absolute = year * 12 + number - 1 + offset
    return f"{absolute // 12:04d}-{absolute % 12 + 1:02d}"


def months_between(first: str, second: str) -> int:
    first_year, first_month = (int(part) for part in first.split("-"))
    second_year, second_month = (int(part) for part in second.split("-"))
    return (second_year - first_year) * 12 + second_month - first_month


def required_text(
    value: Any,
    issues: list[dict[str, Any]],
    table: str,
    row: int,
    field: str,
) -> str | None:
    if value is None or not str(value).strip():
        add_issue(issues, "REQUIRED_VALUE", "必填值为空", table=table, row=row, field=field)
        return None
    text = str(value).strip()
    if text.startswith(FORMULA_PREFIXES):
        add_issue(issues, "FORMULA_INJECTION", "文本疑似公式注入", table=table, row=row, field=field)
        return None
    if PHONE_RE.search(text) or ID_RE.search(text):
        add_issue(issues, "DIRECT_IDENTIFIER", "发现手机号或证件号样式的直接标识", table=table, row=row, field=field)
        return None
    return text


def required_number(
    value: Any,
    issues: list[dict[str, Any]],
    table: str,
    row: int,
    field: str,
    *,
    integer: bool = False,
    minimum: float | None = None,
) -> float | int | None:
    if value is None or value == "":
        add_issue(issues, "REQUIRED_VALUE", "必填值为空", table=table, row=row, field=field)
        return None
    if isinstance(value, bool):
        add_issue(issues, "INVALID_NUMBER", "布尔值不能作为数值", table=table, row=row, field=field)
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        add_issue(issues, "INVALID_NUMBER", "必须是数值", table=table, row=row, field=field)
        return None
    if not math.isfinite(number):
        add_issue(issues, "INVALID_NUMBER", "数值必须有限", table=table, row=row, field=field)
        return None
    if integer and not number.is_integer():
        add_issue(issues, "INVALID_INTEGER", "必须是整数", table=table, row=row, field=field)
        return None
    if minimum is not None and number < minimum:
        add_issue(issues, "NUMBER_OUT_OF_RANGE", f"数值不得小于 {minimum}", table=table, row=row, field=field)
        return None
    return int(number) if integer else number


def required_month(
    value: Any,
    issues: list[dict[str, Any]],
    table: str,
    row: int,
    field: str,
) -> str | None:
    parsed = month_value(value)
    if parsed is None:
        add_issue(issues, "INVALID_MONTH", "月份必须为 Excel 日期、YYYY-MM 或 YYYY-MM-DD", table=table, row=row, field=field)
    return parsed


def required_boolean(
    value: Any,
    issues: list[dict[str, Any]],
    table: str,
    row: int,
    field: str,
) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    text = str(value).strip().lower() if value is not None else ""
    if text in {"是", "可用", "true", "yes", "1"}:
        return True
    if text in {"否", "不可用", "false", "no", "0"}:
        return False
    add_issue(issues, "INVALID_BOOLEAN", "是否可用必须为是/否或布尔值", table=table, row=row, field=field)
    return None


def load_config(path: Path) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    required = {"schema_version", "template_version", "planning", "scenarios", "source"}
    missing = required - set(config)
    if missing:
        raise ValueError(f"配置缺少字段: {', '.join(sorted(missing))}")
    if "基准" not in config["scenarios"]:
        raise ValueError("配置必须包含基准情景")
    if int(config["planning"].get("horizon_months", 0)) != 12:
        raise ValueError("首期 horizon_months 必须为 12")
    return config


def workbook_security_issues(path: Path) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(path) as archive:
            external = [name for name in archive.namelist() if name.startswith("xl/externalLinks/")]
            if external:
                add_issue(issues, "EXTERNAL_LINK", "工作簿包含外部链接")
    except zipfile.BadZipFile:
        add_issue(issues, "INVALID_XLSX", "文件不是有效的 xlsx 工作簿")
        return issues

    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f":
                        add_issue(
                            issues,
                            "EXCEL_FORMULA",
                            "输入工作簿不得包含公式",
                            table=sheet.title,
                            row=cell.row,
                            field=cell.column_letter,
                        )
                    elif isinstance(value, str) and value.strip().startswith(FORMULA_PREFIXES):
                        add_issue(
                            issues,
                            "FORMULA_INJECTION",
                            "文本疑似公式注入",
                            table=sheet.title,
                            row=cell.row,
                            field=cell.column_letter,
                        )
    finally:
        workbook.close()
    return issues


def read_raw_inputs(
    path: Path, config: dict[str, Any], issues: list[dict[str, Any]]
) -> tuple[dict[str, Any], dict[str, list[dict[str, Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    metadata: dict[str, Any] = {}
    tables: dict[str, list[dict[str, Any]]] = {}
    try:
        metadata_spec = config["source"]["metadata"]
        metadata_sheet = metadata_spec["sheet"]
        if metadata_sheet not in workbook.sheetnames:
            add_issue(issues, "MISSING_SHEET", f"缺少工作表 {metadata_sheet}", table=metadata_sheet)
        else:
            sheet = workbook[metadata_sheet]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                add_issue(issues, "EMPTY_SHEET", "元数据表为空", table=metadata_sheet)
            else:
                headers = [str(value).strip() if value is not None else "" for value in rows[0]]
                key_name, value_name = metadata_spec["key_column"], metadata_spec["value_column"]
                if key_name not in headers or value_name not in headers:
                    add_issue(issues, "MISSING_COLUMN", "元数据表缺少字段/值列", table=metadata_sheet)
                else:
                    key_index, value_index = headers.index(key_name), headers.index(value_name)
                    for row in rows[1:]:
                        key = row[key_index] if key_index < len(row) else None
                        value = row[value_index] if value_index < len(row) else None
                        if key is not None and str(key).strip():
                            metadata[str(key).strip()] = value

        for logical_name, spec in config["source"]["tables"].items():
            sheet_name = spec["sheet"]
            tables[logical_name] = []
            if sheet_name not in workbook.sheetnames:
                add_issue(issues, "MISSING_SHEET", f"缺少工作表 {sheet_name}", table=sheet_name)
                continue
            sheet = workbook[sheet_name]
            row_iterator = sheet.iter_rows(values_only=True)
            headers_row = next(row_iterator, None)
            if headers_row is None:
                add_issue(issues, "EMPTY_SHEET", "工作表为空", table=sheet_name)
                continue
            headers = [str(value).strip() if value is not None else "" for value in headers_row]
            missing_columns = [source for source in spec["columns"].values() if source not in headers]
            for column in missing_columns:
                add_issue(issues, "MISSING_COLUMN", f"缺少列 {column}", table=sheet_name, field=column)
            if missing_columns:
                continue
            indexes = {canonical: headers.index(source) for canonical, source in spec["columns"].items()}
            for source_row, row in enumerate(row_iterator, start=2):
                if all(value is None or str(value).strip() == "" for value in row):
                    continue
                record = {
                    canonical: row[index] if index < len(row) else None
                    for canonical, index in indexes.items()
                }
                record["_source_row"] = source_row
                tables[logical_name].append(record)
    finally:
        workbook.close()
    return metadata, tables


def normalize_tables(
    raw: dict[str, list[dict[str, Any]]], issues: list[dict[str, Any]]
) -> dict[str, list[dict[str, Any]]]:
    normalized: dict[str, list[dict[str, Any]]] = {name: [] for name in raw}

    for record in raw.get("enrolled_students", []):
        row = record["_source_row"]
        item = {
            "student_id": required_text(record["student_id"], issues, "在读学员", row, "学员标识"),
            "campus": required_text(record["campus"], issues, "在读学员", row, "校区"),
            "class_type": required_text(record["class_type"], issues, "在读学员", row, "班型"),
            "time_slot": required_text(record["time_slot"], issues, "在读学员", row, "峰值时段"),
            "renewal_month": required_month(record["renewal_month"], issues, "在读学员", row, "续费月份"),
            "status": required_text(record["status"], issues, "在读学员", row, "状态"),
            "source_row": row,
        }
        if item["status"] not in {None, "在读", "停课"}:
            add_issue(issues, "INVALID_STATUS", "状态只接受在读或停课", table="在读学员", row=row, field="状态")
        normalized["enrolled_students"].append(item)

    for record in raw.get("renewal_rates", []):
        row = record["_source_row"]
        rate = required_number(record["base_rate"], issues, "续费率", row, "基准续费率", minimum=0)
        if rate is not None and rate > 1:
            add_issue(issues, "RATE_OUT_OF_RANGE", "续费率必须位于 0 到 1", table="续费率", row=row, field="基准续费率")
        normalized["renewal_rates"].append(
            {
                "campus": required_text(record["campus"], issues, "续费率", row, "校区"),
                "class_type": required_text(record["class_type"], issues, "续费率", row, "班型"),
                "renewal_month": required_month(record["renewal_month"], issues, "续费率", row, "续费月份"),
                "base_rate": rate,
                "source_row": row,
            }
        )

    for record in raw.get("enrollment_plan", []):
        row = record["_source_row"]
        normalized["enrollment_plan"].append(
            {
                "campus": required_text(record["campus"], issues, "招生计划", row, "校区"),
                "class_type": required_text(record["class_type"], issues, "招生计划", row, "班型"),
                "recruitment_month": required_month(record["recruitment_month"], issues, "招生计划", row, "招生月份"),
                "current_target": required_number(record["current_target"], issues, "招生计划", row, "本期目标人数", minimum=0),
                "history_target": required_number(record["history_target"], issues, "招生计划", row, "往期目标人数", minimum=0),
                "history_actual": required_number(record["history_actual"], issues, "招生计划", row, "往期实际人数", minimum=0),
                "start_month": required_month(record["start_month"], issues, "招生计划", row, "开课月份"),
                "time_slot": required_text(record["time_slot"], issues, "招生计划", row, "峰值时段"),
                "source_row": row,
            }
        )
        target = normalized["enrollment_plan"][-1]["history_target"]
        if target == 0:
            add_issue(issues, "ZERO_HISTORY_TARGET", "往期目标必须大于零，无法计算达成率", table="招生计划", row=row, field="往期目标人数")

    for record in raw.get("class_rules", []):
        row = record["_source_row"]
        item = {
            "class_type": required_text(record["class_type"], issues, "班型规则", row, "班型"),
            "target_class_size": required_number(record["target_class_size"], issues, "班型规则", row, "目标班容", integer=True, minimum=1),
            "min_class_size": required_number(record["min_class_size"], issues, "班型规则", row, "最小班容", integer=True, minimum=1),
            "max_class_size": required_number(record["max_class_size"], issues, "班型规则", row, "最大班容", integer=True, minimum=1),
            "teachers_per_class": required_number(record["teachers_per_class"], issues, "班型规则", row, "每班教师数", integer=True, minimum=1),
            "rooms_per_class": required_number(record["rooms_per_class"], issues, "班型规则", row, "每班教室数", integer=True, minimum=1),
            "room_type": required_text(record["room_type"], issues, "班型规则", row, "教室类型"),
            "source_row": row,
        }
        sizes = (item["min_class_size"], item["target_class_size"], item["max_class_size"])
        if all(value is not None for value in sizes) and not sizes[0] <= sizes[1] <= sizes[2]:
            add_issue(issues, "INVALID_CLASS_SIZE_ORDER", "必须满足最小班容 <= 目标班容 <= 最大班容", table="班型规则", row=row)
        normalized["class_rules"].append(item)

    for record in raw.get("teacher_supply", []):
        row = record["_source_row"]
        normalized["teacher_supply"].append(
            {
                "teacher_id": required_text(record["teacher_id"], issues, "教师供给", row, "教师标识"),
                "campus": required_text(record["campus"], issues, "教师供给", row, "校区"),
                "class_type": required_text(record["class_type"], issues, "教师供给", row, "班型"),
                "month": required_month(record["month"], issues, "教师供给", row, "月份"),
                "time_slot": required_text(record["time_slot"], issues, "教师供给", row, "峰值时段"),
                "max_concurrent_classes": required_number(
                    record["max_concurrent_classes"], issues, "教师供给", row, "并发班级上限", minimum=0
                ),
                "source_row": row,
            }
        )

    for record in raw.get("room_supply", []):
        row = record["_source_row"]
        normalized["room_supply"].append(
            {
                "room_id": required_text(record["room_id"], issues, "教室供给", row, "教室标识"),
                "campus": required_text(record["campus"], issues, "教室供给", row, "校区"),
                "room_type": required_text(record["room_type"], issues, "教室供给", row, "教室类型"),
                "month": required_month(record["month"], issues, "教室供给", row, "月份"),
                "time_slot": required_text(record["time_slot"], issues, "教室供给", row, "峰值时段"),
                "available": required_boolean(record["available"], issues, "教室供给", row, "是否可用"),
                "source_row": row,
            }
        )
    return normalized


def duplicate_check(
    records: Iterable[dict[str, Any]],
    fields: tuple[str, ...],
    issues: list[dict[str, Any]],
    table: str,
    code: str,
) -> None:
    seen: dict[tuple[Any, ...], int] = {}
    for record in records:
        key = tuple(record.get(field) for field in fields)
        if any(value is None for value in key):
            continue
        if key in seen:
            add_issue(
                issues,
                code,
                f"唯一键重复，首次出现在第 {seen[key]} 行",
                table=table,
                row=record["source_row"],
                field=" + ".join(fields),
            )
        else:
            seen[key] = record["source_row"]


def validate_relationships(
    tables: dict[str, list[dict[str, Any]]], issues: list[dict[str, Any]]
) -> None:
    duplicate_check(tables["enrolled_students"], ("student_id",), issues, "在读学员", "DUPLICATE_STUDENT")
    duplicate_check(
        tables["renewal_rates"], ("campus", "class_type", "renewal_month"), issues, "续费率", "DUPLICATE_RENEWAL_RATE"
    )
    duplicate_check(
        tables["enrollment_plan"],
        ("campus", "class_type", "start_month", "time_slot"),
        issues,
        "招生计划",
        "DUPLICATE_ENROLLMENT_PLAN",
    )
    duplicate_check(tables["class_rules"], ("class_type",), issues, "班型规则", "DUPLICATE_CLASS_RULE")
    duplicate_check(
        tables["teacher_supply"],
        ("teacher_id", "month", "time_slot"),
        issues,
        "教师供给",
        "TEACHER_TIME_CONFLICT",
    )
    duplicate_check(
        tables["room_supply"],
        ("room_id", "month", "time_slot"),
        issues,
        "教室供给",
        "ROOM_TIME_CONFLICT",
    )

    class_types = {record["class_type"] for record in tables["class_rules"] if record["class_type"]}
    room_types = {record["room_type"] for record in tables["room_supply"] if record["room_type"]}
    renewal_keys = {
        (record["campus"], record["class_type"], record["renewal_month"])
        for record in tables["renewal_rates"]
        if all(record.get(field) is not None for field in ("campus", "class_type", "renewal_month"))
    }
    for table_name in ("enrolled_students", "enrollment_plan", "teacher_supply"):
        for record in tables[table_name]:
            if record.get("class_type") and record["class_type"] not in class_types:
                add_issue(
                    issues,
                    "UNKNOWN_CLASS_TYPE",
                    "班型未在班型规则中定义",
                    table=table_name,
                    row=record["source_row"],
                    field="class_type",
                )
    for record in tables["enrolled_students"]:
        if record.get("status") != "在读":
            continue
        key = (record.get("campus"), record.get("class_type"), record.get("renewal_month"))
        if None not in key and key not in renewal_keys:
            add_issue(
                issues,
                "MISSING_RENEWAL_RATE",
                "在读学员找不到对应续费率",
                table="在读学员",
                row=record["source_row"],
            )
    for record in tables["enrollment_plan"]:
        recruitment_month = record.get("recruitment_month")
        start_month = record.get("start_month")
        if recruitment_month and start_month and start_month < recruitment_month:
            add_issue(
                issues,
                "DATE_ORDER_CONFLICT",
                "开课月份不得早于招生月份",
                table="招生计划",
                row=record["source_row"],
                field="招生月份 + 开课月份",
            )
    for record in tables["class_rules"]:
        if record.get("room_type") and record["room_type"] not in room_types:
            add_issue(
                issues,
                "MISSING_ROOM_TYPE_SUPPLY",
                "班型所需教室类型未出现在教室供给中",
                table="班型规则",
                row=record["source_row"],
                field="教室类型",
            )


def plan_classes(projected_students: float, rule: dict[str, Any]) -> tuple[int, float, str]:
    if projected_students <= 0:
        return 0, 0.0, "无需求"
    if projected_students < rule["min_class_size"]:
        return 0, 0.0, "低于开班下限"

    minimum_classes = math.ceil(projected_students / rule["max_class_size"])
    maximum_classes = math.floor(projected_students / rule["min_class_size"])
    target_classes = math.ceil(projected_students / rule["target_class_size"])
    if minimum_classes <= maximum_classes:
        classes = min(max(target_classes, minimum_classes), maximum_classes)
        status = "可开班"
    else:
        classes = minimum_classes
        status = "末班低于下限"
    average_size = round(projected_students / classes, 4)
    return classes, average_size, status


def build_forecast(
    tables: dict[str, list[dict[str, Any]]], config: dict[str, Any], start_month: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    months = [add_month(start_month, offset) for offset in range(config["planning"]["horizon_months"])]
    rules = {record["class_type"]: record for record in tables["class_rules"]}
    rates = {
        (record["campus"], record["class_type"], record["renewal_month"]): record["base_rate"]
        for record in tables["renewal_rates"]
    }
    students = [record for record in tables["enrolled_students"] if record["status"] == "在读"]
    groups = {
        (record["campus"], record["class_type"], record["time_slot"])
        for record in students + tables["enrollment_plan"]
    }
    teacher_capacity: dict[tuple[str, str, str, str], float] = defaultdict(float)
    for record in tables["teacher_supply"]:
        teacher_capacity[(record["campus"], record["class_type"], record["month"], record["time_slot"])] += record[
            "max_concurrent_classes"
        ]
    room_capacity: dict[tuple[str, str, str, str], int] = defaultdict(int)
    for record in tables["room_supply"]:
        if record["available"]:
            room_capacity[(record["campus"], record["room_type"], record["month"], record["time_slot"])] += 1

    forecast: list[dict[str, Any]] = []
    teacher_gaps: list[dict[str, Any]] = []
    room_demands: dict[tuple[str, str, str, str, str], int] = defaultdict(int)

    for scenario_name, scenario in config["scenarios"].items():
        for month in months:
            for campus, class_type, time_slot in sorted(groups):
                retained = 0.0
                for student in students:
                    if (student["campus"], student["class_type"], student["time_slot"]) != (
                        campus,
                        class_type,
                        time_slot,
                    ):
                        continue
                    if month < student["renewal_month"]:
                        retained += 1.0
                    else:
                        base_rate = rates[(campus, class_type, student["renewal_month"])]
                        retained += min(1.0, base_rate * scenario["renewal_multiplier"])
                new_students = 0.0
                for plan in tables["enrollment_plan"]:
                    if (plan["campus"], plan["class_type"], plan["time_slot"]) != (
                        campus,
                        class_type,
                        time_slot,
                    ) or month < plan["start_month"]:
                        continue
                    attainment = plan["history_actual"] / plan["history_target"]
                    new_students += plan["current_target"] * attainment * scenario["enrollment_multiplier"]
                projected = round(retained + new_students, 4)
                rule = rules[class_type]
                classes, average_class_size, opening_status = plan_classes(projected, rule)
                teacher_demand = classes * rule["teachers_per_class"]
                teacher_supply = teacher_capacity.get((campus, class_type, month, time_slot), 0.0)
                teacher_gap = max(0.0, teacher_demand - teacher_supply)
                evidence = f"FORECAST|{scenario_name}|{campus}|{class_type}|{time_slot}|{month}"
                forecast.append(
                    {
                        "scenario": scenario_name,
                        "month": month,
                        "campus": campus,
                        "class_type": class_type,
                        "time_slot": time_slot,
                        "retained_students": round(retained, 4),
                        "new_students": round(new_students, 4),
                        "projected_students": projected,
                        "planned_classes": classes,
                        "average_class_size": average_class_size,
                        "opening_status": opening_status,
                        "teacher_demand": teacher_demand,
                        "room_type": rule["room_type"],
                        "room_demand": classes * rule["rooms_per_class"],
                        "evidence_key": evidence,
                    }
                )
                teacher_gaps.append(
                    {
                        "scenario": scenario_name,
                        "month": month,
                        "campus": campus,
                        "class_type": class_type,
                        "time_slot": time_slot,
                        "demand": teacher_demand,
                        "capacity": teacher_supply,
                        "gap": teacher_gap,
                        "evidence_key": evidence,
                    }
                )
                room_demands[(scenario_name, campus, rule["room_type"], month, time_slot)] += classes * rule[
                    "rooms_per_class"
                ]

    room_gaps: list[dict[str, Any]] = []
    for (scenario, campus, room_type, month, time_slot), demand in sorted(room_demands.items()):
        capacity = room_capacity.get((campus, room_type, month, time_slot), 0)
        gap = max(0, demand - capacity)
        room_gaps.append(
            {
                "scenario": scenario,
                "month": month,
                "campus": campus,
                "room_type": room_type,
                "time_slot": time_slot,
                "demand": demand,
                "capacity": capacity,
                "gap": gap,
                "evidence_key": f"ROOM|{scenario}|{campus}|{room_type}|{time_slot}|{month}",
            }
        )
    return forecast, teacher_gaps, room_gaps


def positive_runs(records: list[dict[str, Any]], group_fields: tuple[str, ...]) -> list[list[dict[str, Any]]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[tuple(record[field] for field in group_fields)].append(record)
    runs: list[list[dict[str, Any]]] = []
    for group_records in grouped.values():
        current: list[dict[str, Any]] = []
        previous_month: str | None = None
        for record in sorted(group_records, key=lambda item: item["month"]):
            if record["gap"] > 0:
                if current and months_between(previous_month or record["month"], record["month"]) != 1:
                    runs.append(current)
                    current = []
                current.append(record)
                previous_month = record["month"]
            else:
                if current:
                    runs.append(current)
                    current = []
                previous_month = None
        if current:
            runs.append(current)
    return runs


def build_recommendations(
    teacher_gaps: list[dict[str, Any]], room_gaps: list[dict[str, Any]], config: dict[str, Any]
) -> list[dict[str, Any]]:
    planning = config["planning"]
    recommendations: list[dict[str, Any]] = []
    teacher_base = [record for record in teacher_gaps if record["scenario"] == "基准"]
    for run in positive_runs(teacher_base, ("campus", "class_type", "time_slot")):
        if len(run) < planning["teacher_gap_persistence_months"]:
            continue
        first, peak = run[0], max(run, key=lambda item: item["gap"])
        units = math.ceil(peak["gap"])
        evidence = [record["evidence_key"] for record in run]
        recommendations.append(
            {
                "recommendation_id": f"HIRE|{first['campus']}|{first['class_type']}|{first['time_slot']}|{first['month']}",
                "resource_type": "教师",
                "campus": first["campus"],
                "subject": first["class_type"],
                "time_slot": first["time_slot"],
                "first_gap_month": first["month"],
                "peak_gap": units,
                "persistent_months": len(run),
                "lead_months": planning["teacher_lead_months"],
                "action_due_month": add_month(first["month"], -planning["teacher_lead_months"]),
                "action": f"建议在 {add_month(first['month'], -planning['teacher_lead_months'])} 前启动教师容量补充评审，基准情景峰值缺口 {units}。",
                "evidence_keys": evidence,
            }
        )
    room_base = [record for record in room_gaps if record["scenario"] == "基准"]
    for run in positive_runs(room_base, ("campus", "room_type", "time_slot")):
        if len(run) < planning["room_gap_persistence_months"]:
            continue
        first, peak = run[0], max(run, key=lambda item: item["gap"])
        units = math.ceil(peak["gap"])
        evidence = [record["evidence_key"] for record in run]
        recommendations.append(
            {
                "recommendation_id": f"LEASE|{first['campus']}|{first['room_type']}|{first['time_slot']}|{first['month']}",
                "resource_type": "教室",
                "campus": first["campus"],
                "subject": first["room_type"],
                "time_slot": first["time_slot"],
                "first_gap_month": first["month"],
                "peak_gap": units,
                "persistent_months": len(run),
                "lead_months": planning["room_lead_months"],
                "action_due_month": add_month(first["month"], -planning["room_lead_months"]),
                "action": f"建议在 {add_month(first['month'], -planning['room_lead_months'])} 前启动教室扩容方案评审，基准情景峰值缺口 {units}。",
                "evidence_keys": evidence,
            }
        )
    return sorted(recommendations, key=lambda item: (item["action_due_month"], item["resource_type"], item["campus"]))


def scenario_summary(
    teacher_gaps: list[dict[str, Any]], room_gaps: list[dict[str, Any]], scenarios: Iterable[str]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for scenario in scenarios:
        teacher = [record["gap"] for record in teacher_gaps if record["scenario"] == scenario]
        room = [record["gap"] for record in room_gaps if record["scenario"] == scenario]
        rows.append(
            {
                "scenario": scenario,
                "peak_teacher_gap": math.ceil(max(teacher, default=0)),
                "peak_room_gap": math.ceil(max(room, default=0)),
                "teacher_gap_periods": sum(value > 0 for value in teacher),
                "room_gap_periods": sum(value > 0 for value in room),
            }
        )
    return rows


def opening_risks(forecast: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in forecast:
        if record["scenario"] == "基准" and record["opening_status"] not in {"可开班", "无需求"}:
            key = (record["campus"], record["class_type"], record["time_slot"], record["opening_status"])
            grouped[key].append(record)
    risks: list[dict[str, Any]] = []
    for (campus, class_type, time_slot, status), records in sorted(grouped.items()):
        ordered = sorted(records, key=lambda item: item["month"])
        first, last = ordered[0], ordered[-1]
        add_issue(
            risks,
            "CLASS_OPENING_CONSTRAINT",
            f"基准情景 {first['month']} 至 {last['month']} 出现“{status}”，需确认合班、转时段或不开班策略",
            table="预测明细",
            field=f"{campus} / {class_type} / {time_slot}",
            severity=WARNING,
        )
    return risks


def flatten_config(value: Any, prefix: str = "") -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            path = f"{prefix}.{key}" if prefix else key
            rows.extend(flatten_config(value[key], path))
    elif isinstance(value, list):
        rows.append({"path": prefix, "value": json.dumps(value, ensure_ascii=False)})
    else:
        rows.append({"path": prefix, "value": str(value)})
    return rows


def safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def ensure_sheet(workbook: Workbook, name: str):
    return workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)


def write_table(sheet, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    if sheet.max_row >= 4:
        sheet.delete_rows(4, sheet.max_row - 3)
    if not any(
        merged.min_row <= 1 <= merged.max_row and merged.min_col <= 1 <= merged.max_col
        for merged in sheet.merged_cells.ranges
    ):
        try:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        except ValueError:
            pass
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="173F5F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    header_fill = PatternFill("solid", fgColor="20639B")
    thin = Side(style="thin", color="D9E2F3")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row_index, row in enumerate(rows, start=4):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column, value=safe_cell(value))
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 24)
            cell.border = Border(bottom=Side(style="hair", color="E8EEF5"))
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A3:{get_column_letter(max(1, len(headers)))}{max(3, 3 + len(rows))}"
    for column, header in enumerate(headers, start=1):
        values = [header] + [str(row[column - 1]) if row[column - 1] is not None else "" for row in rows]
        width = min(38, max(10, max((len(value) for value in values), default=10) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def rows_from_dicts(records: list[dict[str, Any]], fields: list[tuple[str, str]]) -> tuple[list[str], list[list[Any]]]:
    return [label for _, label in fields], [[record.get(key) for key, _ in fields] for record in records]


def summarize_evidence_keys(keys: list[str]) -> str:
    if not keys:
        return ""
    if len(keys) == 1:
        return keys[0]
    return f"{keys[0]}\n至\n{keys[-1]}\n（共 {len(keys)} 期）"


def write_workbook(
    template: Path,
    destination: Path,
    *,
    status: str,
    metadata: dict[str, Any],
    config: dict[str, Any],
    input_hash: str,
    run_id: str,
    issues: list[dict[str, Any]],
    tables: dict[str, list[dict[str, Any]]],
    forecast: list[dict[str, Any]],
    teacher_gaps: list[dict[str, Any]],
    room_gaps: list[dict[str, Any]],
    recommendations: list[dict[str, Any]],
) -> None:
    workbook = load_workbook(template) if template.exists() else Workbook()
    if workbook.active.title == "Sheet" and len(workbook.sheetnames) == 1:
        workbook.remove(workbook.active)

    summary_rows = [
        ["运行状态", status],
        ["数据日期", metadata.get("data_as_of", "")],
        ["来源系统", metadata.get("source_system", "")],
        ["运行ID", run_id],
        ["输入哈希", input_hash],
        ["配置版本", config["schema_version"]],
        ["模板版本", config["template_version"]],
        ["阻断问题数", sum(issue["severity"] == BLOCKING for issue in issues)],
        ["建议数", len(recommendations)],
        ["人工确认", "必须"],
    ]
    write_table(ensure_sheet(workbook, "运行摘要"), "校区产能规划运行摘要", ["指标", "值"], summary_rows)

    issue_fields = [("severity", "级别"), ("code", "代码"), ("table", "表"), ("row", "行"), ("field", "字段"), ("message", "说明")]
    headers, rows = rows_from_dicts(issues, issue_fields)
    write_table(ensure_sheet(workbook, "校验问题"), "数据校验问题", headers, rows)

    normalized_specs = {
        "标准化在读": ("enrolled_students", [("student_id", "学员标识"), ("campus", "校区"), ("class_type", "班型"), ("time_slot", "峰值时段"), ("renewal_month", "续费月份"), ("status", "状态"), ("source_row", "源行")]),
        "标准化续费": ("renewal_rates", [("campus", "校区"), ("class_type", "班型"), ("renewal_month", "续费月份"), ("base_rate", "基准续费率"), ("source_row", "源行")]),
        "标准化招生": ("enrollment_plan", [("campus", "校区"), ("class_type", "班型"), ("recruitment_month", "招生月份"), ("current_target", "本期目标"), ("history_target", "往期目标"), ("history_actual", "往期实际"), ("start_month", "开课月份"), ("time_slot", "峰值时段"), ("source_row", "源行")]),
        "标准化班型": ("class_rules", [("class_type", "班型"), ("target_class_size", "目标班容"), ("min_class_size", "最小班容"), ("max_class_size", "最大班容"), ("teachers_per_class", "每班教师数"), ("rooms_per_class", "每班教室数"), ("room_type", "教室类型"), ("source_row", "源行")]),
        "标准化教师": ("teacher_supply", [("teacher_id", "教师标识"), ("campus", "校区"), ("class_type", "班型"), ("month", "月份"), ("time_slot", "峰值时段"), ("max_concurrent_classes", "并发班级上限"), ("source_row", "源行")]),
        "标准化教室": ("room_supply", [("room_id", "教室标识"), ("campus", "校区"), ("room_type", "教室类型"), ("month", "月份"), ("time_slot", "峰值时段"), ("available", "是否可用"), ("source_row", "源行")]),
    }
    for sheet_name, (table_name, fields) in normalized_specs.items():
        headers, rows = rows_from_dicts(tables.get(table_name, []), fields)
        write_table(ensure_sheet(workbook, sheet_name), sheet_name, headers, rows)

    forecast_fields = [("scenario", "情景"), ("month", "月份"), ("campus", "校区"), ("class_type", "班型"), ("time_slot", "峰值时段"), ("retained_students", "续费后在读"), ("new_students", "新增学员"), ("projected_students", "预测学员"), ("planned_classes", "计划班级"), ("average_class_size", "平均班额"), ("opening_status", "开班状态"), ("teacher_demand", "教师需求"), ("room_type", "教室类型"), ("room_demand", "教室需求"), ("evidence_key", "证据键")]
    headers, rows = rows_from_dicts(forecast, forecast_fields)
    write_table(ensure_sheet(workbook, "预测明细"), "未来12个月预测明细", headers, rows)

    teacher_fields = [("scenario", "情景"), ("month", "月份"), ("campus", "校区"), ("class_type", "班型"), ("time_slot", "峰值时段"), ("demand", "需求"), ("capacity", "供给"), ("gap", "缺口"), ("evidence_key", "证据键")]
    headers, rows = rows_from_dicts(teacher_gaps, teacher_fields)
    write_table(ensure_sheet(workbook, "教师缺口"), "教师产能缺口", headers, rows)

    room_fields = [("scenario", "情景"), ("month", "月份"), ("campus", "校区"), ("room_type", "教室类型"), ("time_slot", "峰值时段"), ("demand", "需求"), ("capacity", "供给"), ("gap", "缺口"), ("evidence_key", "证据键")]
    headers, rows = rows_from_dicts(room_gaps, room_fields)
    write_table(ensure_sheet(workbook, "教室缺口"), "教室产能缺口", headers, rows)

    recommendation_fields = [("recommendation_id", "建议ID"), ("resource_type", "资源"), ("campus", "校区"), ("subject", "班型/教室类型"), ("time_slot", "峰值时段"), ("first_gap_month", "首次缺口"), ("peak_gap", "峰值缺口"), ("persistent_months", "持续月数"), ("lead_months", "提前期"), ("action_due_month", "建议启动月"), ("action", "建议"), ("evidence_keys", "证据键")]
    recommendation_rows = []
    for record in recommendations:
        normalized = dict(record)
        normalized["evidence_keys"] = summarize_evidence_keys(record["evidence_keys"])
        recommendation_rows.append(normalized)
    headers, rows = rows_from_dicts(recommendation_rows, recommendation_fields)
    write_table(ensure_sheet(workbook, "管理建议"), "管理建议（需人工确认）", headers, rows)

    assumption_rows = [[row["path"], row["value"]] for row in flatten_config(config)]
    write_table(ensure_sheet(workbook, "假设与版本"), "配置、假设与版本", ["路径", "值"], assumption_rows)
    workbook.properties.creator = "xdf-plan-campus-capacity"
    workbook.properties.title = "校区产能规划结果"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()


def write_management_report(
    destination: Path,
    *,
    status: str,
    metadata: dict[str, Any],
    config: dict[str, Any],
    run_id: str,
    input_hash: str,
    issues: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
    recommendations: list[dict[str, Any]],
) -> None:
    lines = [
        "# 校区产能规划管理建议",
        "",
        f"- 运行状态：{status}",
        f"- 数据日期：{metadata.get('data_as_of', '')}",
        f"- 配置版本：{config['schema_version']}",
        f"- 运行 ID：{run_id}",
        f"- 输入哈希：{input_hash}",
        "- 人工确认：必须",
        "",
    ]
    if status == "blocked":
        lines.extend(["## 已阻断", "", "关键数据或安全校验失败，本次不生成招聘或扩租建议。", "", "## 问题", ""])
        for issue in issues:
            lines.append(f"- [{issue['severity']}] {issue['code']}：{issue['message']}（{issue['table']} {issue['row']} {issue['field']}）")
    else:
        lines.extend(["## 情景摘要", "", "| 情景 | 峰值教师缺口 | 峰值教室缺口 | 教师缺口期间数 | 教室缺口期间数 |", "|---|---:|---:|---:|---:|"])
        for summary in summaries:
            lines.append(
                f"| {summary['scenario']} | {summary['peak_teacher_gap']} | {summary['peak_room_gap']} | {summary['teacher_gap_periods']} | {summary['room_gap_periods']} |"
            )
        lines.extend(["", "## 建议", ""])
        if not recommendations:
            lines.append("- 基准情景未达到持续缺口门槛，暂不提出招聘或扩租动作。")
        for recommendation in recommendations:
            lines.extend(
                [
                    f"### {recommendation['recommendation_id']}",
                    "",
                    f"- {recommendation['action']}",
                    f"- 首次缺口：{recommendation['first_gap_month']}；峰值缺口：{recommendation['peak_gap']}；持续：{recommendation['persistent_months']} 个月。",
                    f"- 范围：{recommendation['campus']} / {recommendation['subject']} / {recommendation['time_slot']}。",
                    f"- 证据范围：{summarize_evidence_keys(recommendation['evidence_keys']).replace(chr(10), ' ')}",
                    "",
                ]
            )
        warning_issues = [issue for issue in issues if issue["severity"] == WARNING]
        lines.extend(["## 风险与边界", ""])
        lines.append("- 本结果是容量决策支持，不是具体编制、排课、采购或租赁指令。")
        lines.append("- 发送飞书或钉钉前必须人工预览并确认。")
        for issue in warning_issues:
            lines.append(f"- {issue['code']}：{issue['message']}")
    destination.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def run_planning(
    input_path: Path,
    config_path: Path,
    template_path: Path,
    output_dir: Path,
    *,
    backtest: bool = False,
    reference_date: date | None = None,
) -> dict[str, Any]:
    config = load_config(config_path)
    issues = workbook_security_issues(input_path)
    metadata, raw_tables = read_raw_inputs(input_path, config, issues)
    tables = normalize_tables(raw_tables, issues)
    validate_relationships(tables, issues)

    data_as_of = date_value(metadata.get("data_as_of"))
    if data_as_of is None:
        add_issue(issues, "INVALID_DATA_AS_OF", "元数据 data_as_of 必须为 YYYY-MM-DD", table="元数据", field="data_as_of")
    if not metadata.get("source_system"):
        add_issue(issues, "MISSING_SOURCE_SYSTEM", "元数据缺少 source_system", table="元数据", field="source_system")
    if str(metadata.get("template_version", "")) != config["template_version"]:
        add_issue(issues, "TEMPLATE_VERSION_MISMATCH", "输入模板版本与配置不一致", table="元数据", field="template_version")
    if data_as_of and not backtest:
        today = reference_date or date.today()
        age = (today - data_as_of).days
        if age < -1:
            add_issue(issues, "FUTURE_DATA_DATE", "数据日期晚于运行日期", table="元数据", field="data_as_of")
        elif age > config["planning"]["max_data_age_days"]:
            add_issue(issues, "STALE_DATA", f"数据已超过 {config['planning']['max_data_age_days']} 天", table="元数据", field="data_as_of")

    input_hash = hashlib.sha256(input_path.read_bytes() + config_path.read_bytes()).hexdigest()
    run_id = input_hash[:16]
    blocked = any(issue["severity"] == BLOCKING for issue in issues)
    forecast: list[dict[str, Any]] = []
    teacher_gaps: list[dict[str, Any]] = []
    room_gaps: list[dict[str, Any]] = []
    recommendations: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    if not blocked and data_as_of:
        forecast, teacher_gaps, room_gaps = build_forecast(tables, config, data_as_of.strftime("%Y-%m"))
        issues.extend(opening_risks(forecast))
        recommendations = build_recommendations(teacher_gaps, room_gaps, config)
        summaries = scenario_summary(teacher_gaps, room_gaps, config["scenarios"].keys())

    status = "blocked" if blocked else "ready"
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "campus-capacity-result.xlsx"
    report_path = output_dir / "management-recommendations.md"
    payload_path = output_dir / "message-payload.json"
    metadata_for_output = dict(metadata)
    if data_as_of:
        metadata_for_output["data_as_of"] = data_as_of.isoformat()
    write_workbook(
        template_path,
        workbook_path,
        status=status,
        metadata=metadata_for_output,
        config=config,
        input_hash=input_hash,
        run_id=run_id,
        issues=issues,
        tables=tables,
        forecast=forecast,
        teacher_gaps=teacher_gaps,
        room_gaps=room_gaps,
        recommendations=recommendations,
    )
    write_management_report(
        report_path,
        status=status,
        metadata=metadata_for_output,
        config=config,
        run_id=run_id,
        input_hash=input_hash,
        issues=issues,
        summaries=summaries,
        recommendations=recommendations,
    )
    payload = {
        "schema_version": "1.0.0",
        "status": status,
        "run_id": run_id,
        "data_as_of": metadata_for_output.get("data_as_of"),
        "config_version": config["schema_version"],
        "input_hash": input_hash,
        "requires_human_confirmation": True,
        "auto_send": False,
        "summary": summaries,
        "actions": recommendations,
        "risks": [issue for issue in issues if issue["severity"] == WARNING],
        "blocking_issues": [issue for issue in issues if issue["severity"] == BLOCKING],
        "attachments": [workbook_path.name, report_path.name],
    }
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {
        "status": status,
        "run_id": run_id,
        "issues": issues,
        "forecast": forecast,
        "teacher_gaps": teacher_gaps,
        "room_gaps": room_gaps,
        "recommendations": recommendations,
        "summary": summaries,
        "outputs": {
            "workbook": str(workbook_path),
            "report": str(report_path),
            "payload": str(payload_path),
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="校区教师与教室产能规划")
    parser.add_argument("--input", type=Path, required=True, help="脱敏 xlsx 输入")
    parser.add_argument("--config", type=Path, required=True, help="版本化 JSON 配置")
    parser.add_argument("--template", type=Path, required=True, help="输出 xlsx 模板")
    parser.add_argument("--output-dir", type=Path, required=True, help="输出目录")
    parser.add_argument("--backtest", action="store_true", help="历史回测时跳过数据时效校验")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = run_planning(
            args.input,
            args.config,
            args.template,
            args.output_dir,
            backtest=args.backtest,
        )
    except Exception as exc:  # CLI boundary: return a concise, actionable failure.
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps({"status": result["status"], "run_id": result["run_id"], "outputs": result["outputs"]}, ensure_ascii=False))
    return 2 if result["status"] == "blocked" else 0


if __name__ == "__main__":
    raise SystemExit(main())
